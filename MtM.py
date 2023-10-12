import pandas as pd
import datetime
from openpyxl import load_workbook
import numpy as np


授信机构 = ['法兴中国','法巴中国']

TotalContract = pd.read_excel(r'Z:\LZH\MyTotalContracts.xlsx')
NeedContract1 = TotalContract[TotalContract['OptionType'] == '雪球看涨自动敲出赎回']
NeedContract1 = NeedContract1[NeedContract1['PrePayFee'] <= 0.7]
NeedContract1 = NeedContract1[NeedContract1['LstrikeRatio'] == 0]

NeedContract2 = TotalContract[TotalContract['OptionType'] == '安全气囊']
for Counterparty in 授信机构:
    NeedContract2 = NeedContract2[NeedContract2['CounterParty'] != Counterparty]



NeedContract3 = TotalContract[TotalContract['OptionType'] == '区间保护']
NeedContract3 = NeedContract3[NeedContract3['LstrikeRatio'] == 0]
# NeedContract3 = NeedContract3[NeedContract3['PrePayFee'] <= 0.1]

TotalNeedContract = pd.DataFrame()
TotalNeedContract = TotalNeedContract.append(NeedContract1)
TotalNeedContract = TotalNeedContract.append(NeedContract2)
TotalNeedContract = TotalNeedContract.append(NeedContract3)

ALLNeedContract = TotalNeedContract

TotalNeedContract = TotalNeedContract[TotalNeedContract.BeginDate ==datetime.datetime(2021,2,22)]

filePath = r'C:\Users\Dell\Desktop\测试\雪球组合盯市.xlsx'
df = pd.read_excel(r'C:\Users\Dell\Desktop\测试\雪球组合盯市.xlsx',sheet_name='交易簿记')

class Write_excel():
    def __init__(self, filename):
        self.filename = filename
        self.wb = load_workbook(self.filename)
        self.sheets_name_exist = self.wb.get_sheet_names()
        self.ALLNeedContract = ALLNeedContract

    def write(self, row_n, col_n, value, sheet_name):
        index_temp = self.sheets_name_exist.index(sheet_name)
        self.sheets_temp = self.wb.get_sheet_by_name(self.sheets_name_exist[index_temp])
        self.sheets_temp.cell(row_n, col_n).value = value
        self.wb.save(self.filename)

    def writeTradeBook(self):
        df = pd.read_excel(r'C:\Users\Dell\Desktop\测试\雪球组合盯市.xlsx', sheet_name='交易簿记')
        startindex = df.shape[0] + 2
        index_temp = self.sheets_name_exist.index('交易簿记')
        sheets_temp = self.wb.get_sheet_by_name(self.sheets_name_exist[index_temp])
        for row in range(TotalNeedContract.shape[0]):
            sheets_temp.cell(startindex + row, 1).value = TotalNeedContract.ContractId.values[row]
            sheets_temp.cell(startindex + row, 2).value = TotalNeedContract.CounterParty.values[row]
            sheets_temp.cell(startindex + row, 3).value = TotalNeedContract.Notamt.values[row]
            sheets_temp.cell(startindex + row, 4).value = pd.to_datetime(str(TotalNeedContract.BeginDate.values[row])).date()
            sheets_temp.cell(startindex + row, 5).value = pd.to_datetime(str(TotalNeedContract.EndDate.values[row])).date()
            sheets_temp.cell(startindex + row, 6).value = TotalNeedContract.Code.values[row]
            sheets_temp.cell(startindex + row, 7).value = TotalNeedContract.InitialS.values[row]
            sheets_temp.cell(startindex + row, 8).value = TotalNeedContract.PrePayFee.values[row]
            sheets_temp.cell(startindex + row, 9).value = '=C{0}*H{0}'.format(row+startindex)
            sheets_temp.cell(startindex + row, 10).value = TotalNeedContract.Rebate4.values[row]
            sheets_temp.cell(startindex + row, 11).value = '=C{0}*J{0}'.format(row+startindex)
            sheets_temp.cell(startindex + row, 12).value = '=K{0}'.format(row+startindex)
            sheets_temp.cell(startindex + row, 13).value = '=K{0}'.format(row+startindex)
            sheets_temp.cell(startindex + row, 14).value = '=K{0}'.format(row+startindex)
            sheets_temp.cell(startindex + row, 15).value = 0.00
            sheets_temp.cell(startindex + row, 16).value = pd.to_datetime(str(TotalNeedContract.ActualEndDate.values[row])).date()
            sheets_temp.cell(startindex + row, 17).value = '=IF(E{0}=P{0},IF(P{0}<=TODAY(),1,0),1)'.format(row+startindex)
            sheets_temp.cell(startindex + row, 18).value = '=C{0}*MAX(0,(G{0}-VLOOKUP($W$1,标的收盘价!A:CS,VLOOKUP(F{0},标的簿记!B:C,2,FALSE)+1,FALSE))/G{0})'.format(row+startindex)

        index_temp = self.sheets_name_exist.index('起息结算流水分配')
        sheets_temp = self.wb.get_sheet_by_name(self.sheets_name_exist[index_temp])
        for row in range(TotalNeedContract.shape[0]):
            if TotalNeedContract.OptionType.values[row] == '区间保护':
                sheets_temp.cell(startindex + row, 6).value = pd.to_datetime(str(TotalNeedContract.BeginDate.values[row])).date()
            sheets_temp.cell(startindex + row, 10).value = pd.to_datetime(str(TotalNeedContract.BeginDate.values[row])).date()


        index_temp = self.sheets_name_exist.indeA.x('应收流水')
        sheets_temp = self.wb.get_sheet_by_name(self.sheets_name_exist[index_temp])
        for row in range(TotalNeedContract.shape[0]):
            if TotalNeedContract.OptionType.values[row] == '区间保护':
                sheets_temp.cell(startindex + row, 5).value = round(TotalNeedContract.Notamt.values[row] * TotalNeedContract.TotalPremium.values[row],2)
                sheets_temp.cell(startindex + row, 6).value = pd.to_datetime(str(TotalNeedContract.BeginDate.values[row])).date()
            sheets_temp.cell(startindex + row, 11).value = round(TotalNeedContract.Notamt.values[row] * TotalNeedContract.PrePayFee.values[row],2)
            sheets_temp.cell(startindex + row, 12).value = pd.to_datetime(str(TotalNeedContract.BeginDate.values[row])).date()

        self.wb.save(self.filename)

    # def CheckKnock():

    def BookCashFlow(self):
        df = pd.read_excel(r'C:\Users\Dell\Desktop\测试\雪球组合盯市.xlsx', sheet_name='银行账户')
        df1 = pd.read_excel('E:\myprog\XBY\Settle\CashFlow\\historydetail0222.xlsx', header=1)
        for i in range(len(df1['借方发生额'])):
            df1.loc[i, '借方发生额'] = df1.loc[i, '借方发生额'].replace(',', '')
            df1.loc[i, '贷方发生额'] = df1.loc[i, '贷方发生额'].replace(',', '')
        df1["借方发生额"] = df1["借方发生额"].apply(lambda x: 0 if str(x).isspace() else x).astype(np.float64)
        df1["贷方发生额"] = df1["贷方发生额"].apply(lambda x: 0 if str(x).isspace() else x).astype(np.float64)
        df1['发生额'] = df1['贷方发生额'].sub(df1['借方发生额'])
        # HistoryDetails = df1[['交易时间', '对方单位名称','对方账号', '发生额']].dropna()
        self.HistoryDetails = df1[['交易时间', '对方单位名称','对方账号', '发生额']].dropna()


        loop_value = df.银行账户.values
        self.TodayCashStatement = pd.DataFrame()
        CP = []
        Cash = []
        for i in loop_value:
            if self.HistoryDetails[self.HistoryDetails.对方账号 == i].shape[0] != 0:
                for num in range(self.HistoryDetails[self.HistoryDetails.对方账号 == i].shape[0]):
                    CP.append(df[df.银行账户 == i]['交易对手'].values[0])
                    Cash.append(self.HistoryDetails[self.HistoryDetails.对方账号 == i].发生额.values[num])

        self.TodayCashStatement['交易对手'] = CP
        self.TodayCashStatement['发生额'] = Cash

        # for

        index_temp = self.sheets_name_exist.index('流水簿记')
        sheets_temp = self.wb.get_sheet_by_name(self.sheets_name_exist[index_temp])
        startindex = sheets_temp.max_row + 1

        for row in range(self.TodayCashStatement.shape[0]):
            sheets_temp.cell(startindex + row, 1).value = self.TodayCashStatement.交易对手.values[row]
            sheets_temp.cell(startindex + row, 2).value = self.TodayCashStatement.发生额.values[row]
            sheets_temp.cell(startindex + row, 3).value = datetime.datetime.today().date()

        self.wb.save(self.filename)








































a = Write_excel(filePath)
a.BookCashFlow()
# a.writeTradeBook()





