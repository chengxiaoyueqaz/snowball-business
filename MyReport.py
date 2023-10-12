# -*- coding: utf-8 -*-
"""
Created on Wed Mar 22 09:43:54 2017

@author: User
"""

import numpy as np
import pandas as pd
import os
import openpyxl
import warnings
from sqlalchemy import create_engine
warnings.filterwarnings("ignore")

# import Pricing.optionfunction as of


from WindPy import *

w.start()

from MyReport_Global import ReportInfo
# from Price_Input import AU_Price_Daily
from TradeDate import DealTradeDate
from config import MysqlConfig
YearDays = ReportInfo.YearDays
# OptionDict = KDB_Global.ReportOptionDict
OptionDict = ReportInfo.OptionDict
FutureCode = ReportInfo.FutureCode

ParaTable = ReportInfo.ParaTable
FuturePLCode = ReportInfo.FuturePLCode
Basket = ReportInfo.Basket
RiskTableTitle = ReportInfo.RiskTableTitle
TradeTableTitle = ReportInfo.TradeTableTitle
NewTableTitle = ReportInfo.NewTableTitle
AmericanType = ReportInfo.AmericanType
DepositRate = ReportInfo.DepositRate
CorrRiskTableTitle = ReportInfo.CorrRiskTableTitle
CorrTradeTableTitle = ReportInfo.CorrTradeTableTitle
CorrNewTableTitle = ReportInfo.CorrNewTableTitle
CorrOptionDict = ReportInfo.CorrOptionDict
CorrProduct = ReportInfo.CorrProduct
AU9999P = AU_Price_Daily.AU_1530
AU9999Pre = AU_Price_Daily.AU_Pre_Close



# def DealTradeDate.timedate(datet):
#     return datetime(datet.year, datet.month, datet.day)


class ExoticReport():
    def __init__(self, underlyingcode, Maturity, NotBias=False):
        self.NotBias = NotBias
        self.UNCode = underlyingcode
        self.TDate_today = DealTradeDate.timedate(w.tdaysoffset(0, Maturity).Data[0][0])
        self.TDate_yesterday = DealTradeDate.timedate(w.tdaysoffset(-1, self.TDate_today).Data[0][0])
        self.TDate_before_yesterday = DealTradeDate.timedate(w.tdaysoffset(-1, self.TDate_yesterday).Data[0][0])
        if self.UNCode[0].isdigit():  # 股票标的前面加个stock
            self.filename = (os.getcwd() + '\\ProfitLossAnalyst\\'
                             + self.TDate_today.strftime('%Y%m%d') + 'Stock' + self.UNCode + 'PL.xlsx')
        else:
            self.filename = (os.getcwd() + '\\ProfitLossAnalyst\\'
                             + self.TDate_today.strftime('%Y%m%d') + self.UNCode + 'PL.xlsx')

        self.infoname = ('.\Contracts_and_trade\MyTotalContracts.xlsx')
        # self.infoname=('C:\\Users\\yanghua\\Desktop\\MyTotalContracts.xlsx')
        try:
            totalopinfo = pd.read_excel(self.infoname)
        except:
            raise ValueError('Please check the contract excel whether exist!')

        self.allopinfo = totalopinfo[totalopinfo.Underlying == self.UNCode]
        # TODO A-B美式鲨鱼鳍
        self.allopinfo = self.allopinfo[(self.allopinfo.OptionType == '雪球看涨自动敲出赎回')|(self.allopinfo.OptionType == '逐步调整雪球看涨自动敲出赎回')|
                                        (self.allopinfo.OptionType == '收益凭证雪球')]

        self.opinfo = self.allopinfo[((self.allopinfo.ActualEndDate >= self.TDate_today)
                                      & (totalopinfo.BeginDate < self.TDate_today))]
        self.newinfo = self.allopinfo[self.allopinfo.BeginDate == self.TDate_today]

        # 这里输入Code的昨日和今日的价格，格式'code':(昨日价格：今日价格)
        self.PriceDict = dict()

        if self.UNCode == 'AU':
            # TODO AU9999 20190109收盘价congwind取出来为 None
            # PreAuCash = w.wsd('AU9999.SGE', 'close', self.TDate_yesterday, self.TDate_yesterday).Data[0][0]
            PreAuCash = AU9999Pre
            CurAuCash = (AU9999P if self.TDate_today >= DealTradeDate.timedate(datetime.today()) else
                         w.wsd('AU9999.SGE', 'close', self.TDate_today, self.TDate_today).Data[0][0])  # 这里输入AU9999的今日价格
            self.PriceDict['AU9999.SGE'] = (PreAuCash, CurAuCash)
            # self.PriceDict['AU9999.SGE']=(269.51,269.74)#这里输入AU9999的今日价格
        if self.UNCode in Basket:
            partcode = list(Basket[self.UNCode][:, 0])
            basketweight = np.float64(Basket[self.UNCode][:, 1])
            tempS_yesterday = np.array(w.wsd(partcode, 'close', self.TDate_yesterday, self.TDate_yesterday).Data[0])
            tempS_today = np.array(w.wsd(partcode, 'close', self.TDate_today, self.TDate_today).Data[0])
            tempbasketinfo = self.allopinfo[self.allopinfo.ActualEndDate >= self.TDate_today]
            for basketindex in range(np.size(tempbasketinfo, axis=0)):
                tempbasket = tempbasketinfo.iloc[basketindex, :]
                if tempbasket.Code in self.PriceDict:
                    continue
                else:
                    tempS0 = np.array(w.wsd(partcode, 'close', tempbasket.BeginDate,
                                            tempbasket.BeginDate).Data[0])
                    self.PriceDict[tempbasket.Code] = (
                        np.round((tempS_yesterday * basketweight).dot(1 / tempS0), 6),
                        np.round((tempS_today * basketweight).dot(1 / tempS0), 6))
        else:
            if self.UNCode in FutureCode:
                tempS_yesterday = w.wsd(FutureCode[self.UNCode], 'close', self.TDate_yesterday, self.TDate_yesterday).Data[0][0]
                tempS_today = w.wsd(FutureCode[self.UNCode], 'close', self.TDate_today, self.TDate_today).Data[0][0]
                self.PriceDict[FutureCode[self.UNCode]] = (tempS_yesterday, tempS_today)
            else:
                # tempS_yesterday = w.wsd(self.UNCode,'close',self.TDate_yesterday,self.TDate_yesterday,'PriceAdj=F').Data[0][0]
                tempS_yesterday = w.wsd(self.UNCode, 'close', self.TDate_yesterday, self.TDate_yesterday).Data[0][0]
                tempS_today = w.wsd(self.UNCode, 'close', self.TDate_today, self.TDate_today).Data[0][0]
                self.PriceDict[self.UNCode] = (tempS_yesterday, tempS_today)
        self.FuturePrice_yesterday = tempS_yesterday
        self.FuturePrice_today = tempS_today
        '''
        读取参数表里面的信息
        '''
        """从excel中读参数改成从Mysql数据库中读参数"""
        engine = create_engine(
            "mysql+pymysql://{0}:{1}@{2}/{3}?charset=utf8".format(MysqlConfig.user, MysqlConfig.passwd, MysqlConfig.host, MysqlConfig.databaseName))
        param_today = pd.read_sql('SELECT  underlying,vol_risk, rf_risk,q_risk,vol_t FROM padb.parasetting where recorddate = "{}"'.
                               format(self.TDate_today), engine, index_col='underlying')
        param_yesterday = pd.read_sql('SELECT  underlying,vol_risk, rf_risk,q_risk,vol_t FROM padb.parasetting where recorddate = "{}"'.
                               format(self.TDate_yesterday), engine, index_col='underlying')
        # paraexcel_yesterday = (u'.\每日报价\%s每日报价.xlsx' % self.TDate_yesterday.strftime('%Y%m%d'))
        # paraexcel_today = (u'.\每日报价\%s每日报价.xlsx' % self.TDate_today.strftime('%Y%m%d'))
        # # paraexcel_yesterday = (u'C:\\Users\\yanghua\\Desktop\\每日报价\\'+
        # #                   self.TDate_yesterday.strftime('%Y%m%d')+u'每日报价.xlsx')
        # # paraexcel_today = (u'C:\\Users\\yanghua\\Desktop\\每日报价\\'+
        # #                  self.TDate_today.strftime('%Y%m%d')+u'每日报价.xlsx')
        # para_yesterday = pd.read_excel(paraexcel_yesterday, skiprows=3, usecols='M:W', index_col=0)
        # para_today = pd.read_excel(paraexcel_today, skiprows=3, usecols='M:W', index_col=0)

        paracode = ParaTable[self.UNCode] if self.UNCode in ParaTable else self.UNCode
        self.RiskPara = {'Vol': (param_yesterday['vol_risk'][paracode],
                                 param_today['vol_risk'][paracode]),
                         'Rf': (param_yesterday['rf_risk'][paracode],
                                 param_today['rf_risk'][paracode]),
                         'Dividend': (param_yesterday['q_risk'][paracode],
                                 param_today['q_risk'][paracode])}
        if param_yesterday['rf_risk'][paracode] == DepositRate:
            self.TradePara = {'Vol': (param_yesterday['vol_t'][paracode],
                                      param_today['vol_t'][paracode]),
                              'Rf': (0, 0), 'Dividend': (0, 0)}
        else:
            self.TradePara = {'Vol': (param_yesterday['vol_t'][paracode],
                                      param_today['vol_t'][paracode]),
                              'Rf': (param_yesterday['rf_risk'][paracode],
                                     param_today['rf_risk'][paracode]),
                              'Dividend': (0, 0)}

    def genreport(self):
        writer = pd.ExcelWriter(self.filename, engine='xlsxwriter', datetime_format='yyyy/mm/dd')
        workbook = writer.book
        format0 = workbook.add_format({'font_name': 'Arial Unicode MS',
                                       'num_format': '#,##0.00', 'font_size': 10})
        format1 = workbook.add_format({'font_name': 'Arial Unicode MS',
                                       'num_format': '0.00%', 'font_size': 10})
        format2 = workbook.add_format({'font_name': 'Arial Unicode MS',
                                       'num_format': '#,##0', 'font_size': 10})
        format3 = workbook.add_format({'bold': True, 'font_name': 'Arial Unicode MS',
                                       'num_format': '#,##0', 'font_size': 10,
                                       'bg_color': 'yellow'})
        format4 = workbook.add_format({'bold': True, 'font_name': 'Arial Unicode MS',
                                       'num_format': '_(* #,##0_);_(* (#,##0);_(* "-"_);_(@_)',
                                       'font_size': 10, 'bg_color': 'yellow'})

        if self.opinfo.shape[0] != 0:
            self.GetKnockIn()
            riskgreeks = np.round(self.calgreeks(), 2)
            tradegreeks = np.round(self.calgreeks('Trade'), 2)
            riskgreeks.to_excel(writer, sheet_name=u'风控参数', startcol=4, index=False)
            RiskParaSheet = writer.sheets[u'风控参数']
            RiskParaSheet.set_column('A:AC', 14, format0)
            tradegreeks.to_excel(writer, sheet_name=u'交易参数', startcol=4, index=False)
            TradeParaSheet = writer.sheets[u'交易参数']
            TradeParaSheet.set_column('A:AC', 14, format0)
            self.GetTouchExpire()
        else:
            riskgreeks = pd.DataFrame(pd.Series({'contractid': 'sum', 'code': ''
                                                 }).reindex(index=RiskTableTitle, fill_value=0)).T
            tradegreeks = pd.DataFrame(pd.Series({'contractid': 'sum', 'code': ''
                                                  }).reindex(index=TradeTableTitle, fill_value=0)).T

        if self.newinfo.shape[0] != 0:
            newgreeks = self.newcalgreeks()
            newgreeks.to_excel(writer, sheet_name=u'新起息产品', startcol=4, index=False)
            NewContract = writer.sheets[u'新起息产品']
            NewContract.set_column('A:Y', 14, format0)
        else:
            newgreeks = pd.DataFrame(pd.Series({'contractid': 'sum', 'code': ''
                                                }).reindex(index=NewTableTitle, fill_value=0)).T

        ParaDf = pd.DataFrame({'Info': [u'风控波动率', u'风控无风险利率',
                                        u'风控分红率', u'交易波动率',
                                        u'交易无风险利率', u'交易分红率'],
                               self.TDate_yesterday: [self.RiskPara['Vol'][0], self.RiskPara['Rf'][0],
                                                      self.RiskPara['Dividend'][0],
                                                      self.TradePara['Vol'][0], self.TradePara['Rf'][0],
                                                      self.TradePara['Dividend'][0]],
                               self.TDate_today: [self.RiskPara['Vol'][1], self.RiskPara['Rf'][1],
                                                  self.RiskPara['Dividend'][1],
                                                  self.TradePara['Vol'][1], self.TradePara['Rf'][1],
                                                  self.TradePara['Dividend'][1]]},
                              columns=['Info', self.TDate_yesterday, self.TDate_today])
        PriceDf = pd.DataFrame(self.PriceDict, index=[self.TDate_yesterday, self.TDate_today]).T
        PriceDf['Info'] = PriceDf.index
        PriceDf = PriceDf.reindex(columns=['Info', self.TDate_yesterday, self.TDate_today])
        BasicInfo = ParaDf.append(PriceDf, ignore_index=True)
        BasicInfo.to_excel(writer, sheet_name='Sheet1', startrow=1, index=False)

        self.allopinfo.to_excel(writer, sheet_name='Sheet1', startcol=4, index=False)
        worksheet1 = writer.sheets['Sheet1']
        worksheet1.set_column('A:AZ', 14)
        worksheet1.write_string(0, 1, u'日期1')
        worksheet1.write_string(0, 2, u'日期2')
        worksheet1.conditional_format('B3:C8', {'type': 'cell', 'criteria': '<', 'value': 5, 'format': format1})

        if (self.newinfo.shape[0] != 0) & (self.opinfo.shape[0] != 0):
            RiskParaSheet.conditional_format('B3:C8', {'type': 'cell', 'criteria': '<', 'value': 5, 'format': format1})
            BasicInfo.to_excel(writer, sheet_name=u'风控参数', startrow=1, index=False)
            TradeParaSheet.conditional_format('B3:C8', {'type': 'cell', 'criteria': '<', 'value': 5, 'format': format1})
            BasicInfo.to_excel(writer, sheet_name=u'交易参数', startrow=1, index=False)
            NewContract.conditional_format('B3:C8', {'type': 'cell', 'criteria': '<', 'value': 5, 'format': format1})
            BasicInfo.to_excel(writer, sheet_name=u'新起息产品', startrow=1, index=False)
        elif (self.newinfo.shape[0] != 0) & (self.opinfo.shape[0] == 0):
            NewContract.conditional_format('B3:C8', {'type': 'cell', 'criteria': '<', 'value': 5, 'format': format1})
            BasicInfo.to_excel(writer, sheet_name=u'新起息产品', startrow=1, index=False)
        elif (self.newinfo.shape[0] == 0) & (self.opinfo.shape[0] != 0):
            RiskParaSheet.conditional_format('B3:C8', {'type': 'cell', 'criteria': '<', 'value': 5, 'format': format1})
            BasicInfo.to_excel(writer, sheet_name=u'风控参数', startrow=1, index=False)
            TradeParaSheet.conditional_format('B3:C8', {'type': 'cell', 'criteria': '<', 'value': 5, 'format': format1})
            BasicInfo.to_excel(writer, sheet_name=u'交易参数', startrow=1, index=False)

        PoolTitle = ['Date', 'RiskMonte', 'SSigB', 'SSigT', 'Risk-Trade', 'FutureP&L',
                     'TotalP&L', 'NewP&L', 'TotalExistP&L']
        # TODO change Excel to mysql
        ProfitLoss_Df = pd.read_excel('.\Contracts_and_trade\对冲头寸损益.xlsx',
                                      sheet_name=FuturePLCode[self.UNCode], index_col=0)
        # ProfitLoss_Df = pd.read_excel(u'C:\\Users\\yanghua\\Desktop\\对冲头寸损益.xlsx',
        #                   sheet_name=FuturePLCode[self.UNCode],index_col=0)
        # TODO change Excel to mysql
        print('Get Future PL of %s'%self.UNCode)
        FuturePL = ProfitLoss_Df[u'收盘价当日盈亏'][self.TDate_today]
        FutureDeltaPL = ProfitLoss_Df[u'昨日Delta盈亏'][self.TDate_today]
        FutureGammaPL = FuturePL - FutureDeltaPL
        self.PoolInfo = pd.DataFrame()
        TempInfoSeri = [self.TDate_yesterday, self.TDate_today, 'Profit&Loss', 'Delta1', 'Delta2',
                        'BiasP&L', 'PositionP&L', 'DeltaP&L', 'Gamma1', 'Gamma2',
                        'GammaP&L', 'Theta1', 'Theta2', 'ThetaP&L', 'Gamma+ThetaP&L',
                        'Vega1', 'Vega2', 'VegaP&L', 'GreeksPL', 'GreeksError']
        self.PoolInfo['Date'] = TempInfoSeri
        tempfuturepl = np.array([np.nan, np.nan, FuturePL, np.nan, np.nan, 0, 0, FutureDeltaPL,
                                 np.nan, np.nan, FutureGammaPL, np.nan, np.nan, 0,
                                 FutureGammaPL, np.nan, np.nan, 0, FuturePL, np.nan])
        self.PoolInfo['FutureP&L'] = tempfuturepl
        tempRiskMonte = np.append(np.array(
            [riskgreeks['RiskMonteValue1'].iloc[-1],
             riskgreeks['RiskMonteValue2'].iloc[-1] +newgreeks['RiskMonteValue'].iloc[-1],
             riskgreeks['RiskMontePL'].iloc[-1] +newgreeks['OptionFee'].iloc[-1] -newgreeks['RiskMonteValue'].iloc[-1]]),
            np.nan * np.arange(17))
        self.PoolInfo['RiskMonte'] = tempRiskMonte
        tempRiskCol = np.array([riskgreeks['RiskPrice1'].iloc[-1],
                                riskgreeks['RiskPrice2'].iloc[-1] +newgreeks['RiskPrice'].iloc[-1],
                                riskgreeks['RiskPL'].iloc[-1] +newgreeks['OptionFee'].iloc[-1] +newgreeks['RiskPrice'].iloc[-1],
                                riskgreeks['RiskDelta1'].iloc[-1],
                                riskgreeks['RiskDelta2'].iloc[-1] +newgreeks['RiskDelta'].iloc[-1],
                                riskgreeks['RiskBiasPL'].iloc[-1], 0,
                                riskgreeks['RiskDeltaPL'].iloc[-1],
                                riskgreeks['RiskGamma1'].iloc[-1],
                                riskgreeks['RiskGamma2'].iloc[-1] +newgreeks['RiskGamma'].iloc[-1],
                                riskgreeks['RiskGammaPL'].iloc[-1],
                                riskgreeks['RiskTheta1'].iloc[-1],
                                riskgreeks['RiskTheta2'].iloc[-1] +newgreeks['RiskTheta'].iloc[-1],
                                riskgreeks['RiskThetaPL'].iloc[-1],
                                riskgreeks['RiskGammaPL'].iloc[-1] +riskgreeks['RiskThetaPL'].iloc[-1],
                                riskgreeks['RiskVega1'].iloc[-1],
                                riskgreeks['RiskVega2'].iloc[-1] +newgreeks['RiskVega'].iloc[-1],
                                riskgreeks['RiskVegaPL'].iloc[-1],
                                riskgreeks['RiskGreeksPL'].iloc[-1],
                                riskgreeks['RiskPL'].iloc[-1] -riskgreeks['RiskGreeksPL'].iloc[-1]])
        self.PoolInfo['SSigB'] = tempRiskCol
        tempTradeCol = np.array([tradegreeks['TradePrice1'].iloc[-1],
                                 tradegreeks['TradePrice2'].iloc[-1] +newgreeks['TradePrice'].iloc[-1],
                                 tradegreeks['TradePL'].iloc[-1] +newgreeks['OptionFee'].iloc[-1] +newgreeks['TradePrice'].iloc[-1],
                                 tradegreeks['TradeDelta1'].iloc[-1],
                                 tradegreeks['TradeDelta2'].iloc[-1] +newgreeks['TradeDelta'].iloc[-1],
                                 tradegreeks['TradeBiasPL'].iloc[-1], 0,
                                 tradegreeks['TradeDeltaPL'].iloc[-1],
                                 tradegreeks['TradeGamma1'].iloc[-1],
                                 tradegreeks['TradeGamma2'].iloc[-1] +newgreeks['TradeGamma'].iloc[-1],
                                 tradegreeks['TradeGammaPL'].iloc[-1],
                                 tradegreeks['TradeTheta1'].iloc[-1],
                                 tradegreeks['TradeTheta2'].iloc[-1] +newgreeks['TradeTheta'].iloc[-1],
                                 tradegreeks['TradeThetaPL'].iloc[-1],
                                 tradegreeks['TradeGammaPL'].iloc[-1] +tradegreeks['TradeThetaPL'].iloc[-1],
                                 tradegreeks['TradeVega1'].iloc[-1],
                                 tradegreeks['TradeVega2'].iloc[-1] +newgreeks['TradeVega'].iloc[-1],
                                 tradegreeks['TradeVegaPL'].iloc[-1],
                                 tradegreeks['TradeGreeksPL'].iloc[-1],
                                 tradegreeks['TradePL'].iloc[-1] -tradegreeks['TradeGreeksPL'].iloc[-1]])
        self.PoolInfo['SSigT'] = tempTradeCol
        tempRiskTrade = tempRiskCol - tempTradeCol
        tempRiskTrade[[0, 1, 3, 4, 8, 9, 11, 12, 15, 16, 19]] = np.nan
        self.PoolInfo['Risk-Trade'] = tempRiskTrade
        tempRiskTotalPL = tempfuturepl + tempRiskCol
        tempRiskTotalPL[6] = tempRiskTotalPL[7] - tempRiskTotalPL[5]
        self.PoolInfo['TotalP&L'] = tempRiskTotalPL
        tempNewRisk = np.array([newgreeks['OptionFee'].iloc[-1],
                                newgreeks['RiskPrice'].iloc[-1],
                                newgreeks['OptionFee'].iloc[-1] +newgreeks['RiskPrice'].iloc[-1],
                                np.nan, newgreeks['RiskDelta'].iloc[-1], 0, 0, 0, np.nan,
                                newgreeks['RiskGamma'].iloc[-1], 0, np.nan,
                                newgreeks['RiskTheta'].iloc[-1], 0, 0, np.nan,
                                newgreeks['RiskVega'].iloc[-1], 0,
                                newgreeks['OptionFee'].iloc[-1] +newgreeks['RiskPrice'].iloc[-1], np.nan])
        self.PoolInfo['NewP&L'] = tempNewRisk
        self.PoolInfo['TotalExistP&L'] = tempRiskTotalPL - tempNewRisk
        self.PoolInfo = np.round(self.PoolInfo)
        self.PoolInfo = self.PoolInfo.reindex(columns=PoolTitle)
        self.PoolInfo.to_excel(writer, sheet_name=u'交易盈亏分解', index=False)
        worksheet5 = writer.sheets[u'交易盈亏分解']
        worksheet5.set_column('A:S', 14, format2)
        worksheet5.set_column('F:I', 14, format3)
        worksheet5.conditional_format('H2:H21', {'type': 'cell', 'criteria': '==', 'value': 0, 'format': format4})
        writer.save()
        self.GatherExoticReport()

    def calgreeks(self, ParaType='Risk'):
        TableTitle = RiskTableTitle if ParaType == 'Risk' else TradeTableTitle
        TempPara = self.RiskPara if ParaType == 'Risk' else self.TradePara
        exogreeks = pd.DataFrame()
        for i in range(np.size(self.opinfo, axis=0)):
            singleinfo = self.opinfo.iloc[i, :].copy()
            print("start make ",singleinfo.ContractId)
            if (singleinfo.Code in self.PriceDict):
                S_yesterday = self.PriceDict[singleinfo.Code][0]
                S_today = self.PriceDict[singleinfo.Code][1]
            else:
                S_yesterday = w.wsd(singleinfo.Code, 'close', self.TDate_yesterday, self.TDate_yesterday).Data[0][0]
                S_today = w.wsd(singleinfo.Code, 'close', self.TDate_today, self.TDate_today).Data[0][0]
                self.PriceDict[singleinfo.Code] = (S_yesterday, S_today)
            if singleinfo.SettleMethod == 'TWAP' and singleinfo.ActualEndDate == self.TDate_today:
                S_today = singleinfo.TWAPSettlePrice
            actualnotamt = (singleinfo.Notamt * singleinfo.Maturity /
                            singleinfo.YearDays * singleinfo.Pratio *
                            singleinfo.Pos if singleinfo.IfActual == 0 else
                            singleinfo.Notamt * singleinfo.Pratio *
                            singleinfo.Pos)
            multicontract = actualnotamt / singleinfo.InitialS
            Maturity_yesterday = of.mytdate(self.TDate_yesterday, singleinfo.EndDate)
            Maturity_today = (of.mytdate(self.TDate_today, singleinfo.ActualEndDate)
                  if singleinfo.OptionType in AmericanType else
                  of.mytdate(self.TDate_today, singleinfo.EndDate))
            # 暂时先将MontePrice用Price的定价方法替代
            tempMonte_yesterday = (OptionDict[singleinfo.OptionType](S_yesterday, Maturity_yesterday, TempPara['Vol'][0],
                                                                     TempPara['Rf'][0], TempPara['Dividend'][0],
                                                                     'Price', singleinfo) +
                                   self.AccrualHist(singleinfo, self.TDate_before_yesterday,
                                                    'Price')) * multicontract
            tempMonte_today = (OptionDict[singleinfo.OptionType](S_today, Maturity_today, TempPara['Vol'][1],
                                                                 TempPara['Rf'][1], TempPara['Dividend'][1],
                                                                 'Price', singleinfo) +
                               self.AccrualHist(singleinfo, self.TDate_yesterday, 'Price')) * multicontract
            tempMontePL = np.array([tempMonte_yesterday, tempMonte_today, tempMonte_today - tempMonte_yesterday])
            tempPrice_yesterday = (OptionDict[singleinfo.OptionType](S_yesterday, Maturity_yesterday, TempPara['Vol'][0],
                                                                     TempPara['Rf'][0], TempPara['Dividend'][0],
                                                                     'Price', singleinfo) +
                                   self.AccrualHist(singleinfo, self.TDate_before_yesterday, 'Price')) * multicontract
            tempPrice_today = (OptionDict[singleinfo.OptionType](S_today, Maturity_today, TempPara['Vol'][1],
                                                                 TempPara['Rf'][1], TempPara['Dividend'][1], 'Price',
                                                                 singleinfo) +
                               self.AccrualHist(singleinfo, self.TDate_yesterday, 'Price')) * multicontract
            """指数增强型收益互换每日盈亏需要增加一个应计利息"""
            tempPricePL = np.array([tempPrice_yesterday, tempPrice_today, tempPrice_today - tempPrice_yesterday +
                                    (singleinfo.Notamt/singleinfo.YearDays*singleinfo.Premium if singleinfo.OptionType=='指数增强型收益互换' else 0)])
            tempDelta_yesterday = OptionDict[singleinfo.OptionType](S_yesterday, Maturity_yesterday, TempPara['Vol'][0],
                                                                    TempPara['Rf'][0], TempPara['Dividend'][0], 'Delta',
                                                                    singleinfo) * multicontract
            tempDelta_today = OptionDict[singleinfo.OptionType](S_today, Maturity_today, TempPara['Vol'][1],
                                                                TempPara['Rf'][1], TempPara['Dividend'][1], 'Delta',
                                                                singleinfo) * multicontract
            tempDeltaPL = np.array(
                [tempDelta_yesterday, tempDelta_today, tempDelta_yesterday * (S_today - S_yesterday)])
            tempGamma_yesterday = OptionDict[singleinfo.OptionType](S_yesterday, Maturity_yesterday, TempPara['Vol'][0],
                                                                    TempPara['Rf'][0], TempPara['Dividend'][0], 'Gamma',
                                                                    singleinfo) * multicontract
            tempGamma_today = OptionDict[singleinfo.OptionType](S_today, Maturity_today, TempPara['Vol'][1],
                                                                TempPara['Rf'][1], TempPara['Dividend'][1], 'Gamma',
                                                                singleinfo) * multicontract
            tempGammaPL = np.array([tempGamma_yesterday, tempGamma_today, 0.5 * tempGamma_yesterday * (S_today - S_yesterday) ** 2])
            tempTheta_yesterday = OptionDict[singleinfo.OptionType](S_yesterday, Maturity_yesterday, TempPara['Vol'][0],
                                                                    TempPara['Rf'][0], TempPara['Dividend'][0], 'Theta',
                                                                    singleinfo) * multicontract
            tempTheta_today = OptionDict[singleinfo.OptionType](S_today, Maturity_today, TempPara['Vol'][1],
                                                                TempPara['Rf'][1], TempPara['Dividend'][1], 'Theta',
                                                                singleinfo) * multicontract
            tempThetaPL = np.array([tempTheta_yesterday, tempTheta_today, tempTheta_yesterday / YearDays])
            tempVega_yesterday = OptionDict[singleinfo.OptionType](S_yesterday, Maturity_yesterday, TempPara['Vol'][0],
                                                                   TempPara['Rf'][0], TempPara['Dividend'][0], 'Vega',
                                                                   singleinfo) * multicontract
            tempVega_today = OptionDict[singleinfo.OptionType](S_today, Maturity_today, TempPara['Vol'][1],
                                                               TempPara['Rf'][1], TempPara['Dividend'][1], 'Vega',
                                                               singleinfo) * multicontract
            tempVegaPL = np.array([tempVega_yesterday, tempVega_today,
                                   tempVega_yesterday * (TempPara['Vol'][1] - TempPara['Vol'][0])])
            tempGreeksPL = tempDeltaPL[-1] + tempGammaPL[-1] + tempThetaPL[-1] + tempVegaPL[-1]
            tempVanna = (OptionDict[singleinfo.OptionType](S_today, Maturity_today, TempPara['Vol'][1] + 0.001,
                                                           TempPara['Rf'][1], TempPara['Dividend'][1], 'Delta',
                                                           singleinfo) -
                         OptionDict[singleinfo.OptionType](S_today, Maturity_today, TempPara['Vol'][1] - 0.001,
                                                           TempPara['Rf'][1], TempPara['Dividend'][1], 'Delta',
                                                           singleinfo)) / 0.002 * multicontract
            tempZomma = (OptionDict[singleinfo.OptionType](S_today, Maturity_today, TempPara['Vol'][1] + 0.001,
                                                           TempPara['Rf'][1], TempPara['Dividend'][1], 'Gamma',
                                                           singleinfo) -
                         OptionDict[singleinfo.OptionType](S_today, Maturity_today, TempPara['Vol'][1] - 0.001,
                                                           TempPara['Rf'][1], TempPara['Dividend'][1], 'Gamma',
                                                           singleinfo)) / 0.002 * multicontract
            tempVolga = (OptionDict[singleinfo.OptionType](S_today, Maturity_today, TempPara['Vol'][1] + 0.001,
                                                           TempPara['Rf'][1], TempPara['Dividend'][1], 'Vega',
                                                           singleinfo) -
                         OptionDict[singleinfo.OptionType](S_today, Maturity_today, TempPara['Vol'][1] - 0.001,
                                                           TempPara['Rf'][1], TempPara['Dividend'][1], 'Vega',
                                                           singleinfo)) / 0.002 * multicontract
            if (self.UNCode in Basket) | (self.NotBias) | (self.UNCode not in FutureCode):
                tempbiasPL = 0
            else:
                if (('.SH' in self.UNCode) or ('.SZ' in self.UNCode)) and (FutureCode[self.UNCode] != singleinfo.Code):
                    tempbiasPL = tempDelta_yesterday * (S_today - S_yesterday - (self.FuturePrice_today - self.FuturePrice_yesterday) * S_yesterday / self.FuturePrice_yesterday)
                else:
                    try:
                        tempbiasPL = tempDelta_yesterday * (S_today - S_yesterday - self.FuturePrice_today + self.FuturePrice_yesterday)
                    except:
                        print(self.UNCode)
                        print(singleinfo)
            tempnum = np.concatenate([tempMontePL, tempPricePL, tempDeltaPL,
                                      tempGammaPL, tempThetaPL, tempVegaPL,
                                      [tempGreeksPL, tempbiasPL, tempVanna, tempZomma,
                                       tempVolga]]) * (-1)  # 按照现金流来计算
            tempgreeks = pd.DataFrame(tempnum, index=TableTitle[2:]).T
            tempgreeks[TableTitle[0]] = singleinfo.ContractId
            tempgreeks[TableTitle[1]] = singleinfo.Code
            print(tempnum)
            exogreeks = exogreeks.append(tempgreeks)
            print('end make ',singleinfo.ContractId)
        totaltemp = exogreeks.iloc[:, :-2].sum()
        totaltemp[TableTitle[0]] = 'sum'
        totaltemp[TableTitle[1]] = ''
        exogreeks = exogreeks.append(totaltemp, ignore_index=True)
        exogreeks = exogreeks.reindex(columns=TableTitle)
        return exogreeks

    def newcalgreeks(self):
        newgreeks = pd.DataFrame()
        for j in range(np.size(self.newinfo, axis=0)):
            newsingleinfo = self.newinfo.iloc[j, :].copy()
            if newsingleinfo.Code in self.PriceDict:
                newcash = self.PriceDict[newsingleinfo.Code][1]
            else:
                S_yesterday = w.wsd(newsingleinfo.Code, 'close', self.TDate_yesterday,
                                    self.TDate_yesterday).Data[0][0]
                newcash = w.wsd(newsingleinfo.Code, 'close', self.TDate_today,
                                self.TDate_today).Data[0][0]
                self.PriceDict[newsingleinfo.Code] = (S_yesterday, newcash)
            # 这里Basket产品的价格肯定有
            TotalGreeks = np.concatenate([self.GetGreeks(newcash, newsingleinfo, self.RiskPara),
                                          self.GetGreeks(newcash, newsingleinfo, self.TradePara)]) * (-1)
            tempnewgreeks = pd.DataFrame(TotalGreeks, index=NewTableTitle[3:]).T

            if newsingleinfo.OptionType == u'看涨自动敲出赎回' or newsingleinfo.OptionType == u'看涨自动敲出赎回转看涨' or newsingleinfo.OptionType == u'敲入看涨自动敲出赎回':
                ObDate = newsingleinfo.ObSerDate.split('\n')
                ObDate = list(map(lambda x: datetime.strptime(x, '%Y/%m/%d'), ObDate))
                ObDate = sorted(ObDate)
                BeginDate = DealTradeDate.timedate(newsingleinfo.BeginDate)
                TotalDays = newsingleinfo.Maturity
                SettleChazhi = (DealTradeDate.timedate(newsingleinfo.DealDate) - DealTradeDate.timedate(newsingleinfo.EndDate)).days
                SettleRatio = list(map(lambda x: ((x - BeginDate).days + SettleChazhi) / TotalDays, ObDate))
                multicontract = (newsingleinfo.Notamt * newsingleinfo.Maturity /
                                 newsingleinfo.YearDays * newsingleinfo.Pratio *
                                 newsingleinfo.Pos if newsingleinfo.IfActual == 0 else
                                 newsingleinfo.Notamt * newsingleinfo.Pratio *
                                 newsingleinfo.Pos) / newsingleinfo.InitialS
                tempfee = of.AutoCallFee(newcash, newsingleinfo.InitialS * newsingleinfo.HtouchRatio,
                                         self.TDate_today, self.RiskPara['Vol'][1], self.RiskPara['Rf'][1],
                                         self.RiskPara['Dividend'][1],
                                         newsingleinfo.InitialS * newsingleinfo.Premium / newsingleinfo.Pratio,
                                         ObDate, SettleRatio) * multicontract
            elif newsingleinfo.OptionType == u'看跌自动敲出赎回':
                ObDate = newsingleinfo.ObSerDate.split('\n')
                ObDate = list(map(lambda x: datetime.strptime(x, '%Y/%m/%d'), ObDate))
                ObDate = sorted(ObDate)
                BeginDate = DealTradeDate.timedate(newsingleinfo.BeginDate)
                TotalDays = newsingleinfo.Maturity
                SettleChazhi = (DealTradeDate.timedate(newsingleinfo.DealDate) - DealTradeDate.timedate(newsingleinfo.EndDate)).days
                SettleRatio = list(map(lambda x: ((x - BeginDate).days + SettleChazhi) / TotalDays, ObDate))
                multicontract = (newsingleinfo.Notamt * newsingleinfo.Maturity /
                                 newsingleinfo.YearDays * newsingleinfo.Pratio *
                                 newsingleinfo.Pos if newsingleinfo.IfActual == 0 else
                                 newsingleinfo.Notamt * newsingleinfo.Pratio *
                                 newsingleinfo.Pos) / newsingleinfo.InitialS
                tempfee = of.AutoPutFee(newcash, newsingleinfo.InitialS * newsingleinfo.LtouchRatio,
                                        self.TDate_today, self.RiskPara['Vol'][1], self.RiskPara['Rf'][1],
                                        self.RiskPara['Dividend'][1],
                                        newsingleinfo.InitialS * newsingleinfo.Premium / newsingleinfo.Pratio,
                                        ObDate, SettleRatio) * multicontract
            elif newsingleinfo.OptionType == u'逐步调整看涨自动敲出赎回':
                ObDate = newsingleinfo.ObSerDate.split('\n')
                ObDate = list(map(lambda x: datetime.strptime(x, '%Y/%m/%d'), ObDate))
                ObDate = sorted(ObDate)
                BeginDate = DealTradeDate.timedate(newsingleinfo.BeginDate)
                TotalDays = newsingleinfo.Maturity
                SettleChazhi = (DealTradeDate.timedate(newsingleinfo.DealDate) - DealTradeDate.timedate(newsingleinfo.EndDate)).days
                SettleRatio = list(map(lambda x: ((x - BeginDate).days + SettleChazhi) / TotalDays, ObDate))
                multicontract = (newsingleinfo.Notamt * newsingleinfo.Maturity /
                                 newsingleinfo.YearDays * newsingleinfo.Pratio *
                                 newsingleinfo.Pos if newsingleinfo.IfActual == 0 else
                                 newsingleinfo.Notamt * newsingleinfo.Pratio *
                                 newsingleinfo.Pos) / newsingleinfo.InitialS
                tempfee = of.AutoCallStepChageFee(newcash, newsingleinfo.InitialS * newsingleinfo.HtouchRatio,
                                                  self.TDate_today, self.RiskPara['Vol'][1], self.RiskPara['Rf'][1],
                                                  self.RiskPara['Dividend'][1],
                                                  newsingleinfo.InitialS * newsingleinfo.Rebate3,
                                                  newsingleinfo.InitialS * newsingleinfo.Premium / newsingleinfo.Pratio,
                                                  ObDate, SettleRatio) * multicontract
            # elif newsingleinfo.ContractId=='SWHY-GTJA-18001':
            #
            elif newsingleinfo.OptionType=='指数增强型收益互换':
                tempfee = (self.TDate_today - newsingleinfo.BeginDate).days*newsingleinfo.Notamt * newsingleinfo.Premium/newsingleinfo.YearDays
            elif newsingleinfo.OptionType=='雪球看涨自动敲出赎回' or newsingleinfo.OptionType=='逐步调整雪球看涨自动敲出赎回' or newsingleinfo.OptionType=='收益凭证雪球':
                """雪球看涨自动敲出赎回和逐步调整雪球看涨自动敲出赎回的结构设期权费为0"""
                tempfee = 0
            else:
                tempfee = (newsingleinfo.Notamt * newsingleinfo.Maturity /
                           newsingleinfo.YearDays * newsingleinfo.Premium
                           if newsingleinfo.IfActual == 0 else
                           newsingleinfo.Notamt * newsingleinfo.Premium)
            tempnewgreeks[NewTableTitle[2]] = tempfee
            tempnewgreeks[NewTableTitle[0]] = newsingleinfo.ContractId
            tempnewgreeks[NewTableTitle[1]] = newsingleinfo.Code
            newgreeks = newgreeks.append(tempnewgreeks)
        totaltemp = newgreeks.iloc[:, :-2].sum()
        totaltemp[NewTableTitle[0]] = 'sum'
        totaltemp[NewTableTitle[1]] = ''

        newgreeks = newgreeks.append(totaltemp, ignore_index=True)
        newgreeks = newgreeks.reindex(columns=NewTableTitle)
        return newgreeks

    def GetGreeks(self, S, singleinfo, Para):
        multicontract = (singleinfo.Notamt * singleinfo.Maturity /
                         singleinfo.YearDays * singleinfo.Pratio * singleinfo.Pos
                         if singleinfo.IfActual == 0 else
                         singleinfo.Notamt * singleinfo.Pratio * singleinfo.Pos) / singleinfo.InitialS
        t2 = of.mytdate(self.TDate_today, singleinfo.EndDate)
        if singleinfo.OptionType == u'看涨自动敲出赎回':
            TempOptionType = u'看涨自动敲出赎回新起息'
        elif singleinfo.OptionType == u'看跌自动敲出赎回':
            TempOptionType = u'看跌自动敲出赎回新起息'
        elif singleinfo.OptionType == u'看涨自动敲出赎回转看涨':
            TempOptionType = u'看涨自动敲出赎回转看涨新起息'
        elif singleinfo.OptionType == u'逐步调整看涨自动敲出赎回':
            TempOptionType = u'逐步调整看涨自动敲出赎回新起息'
        elif singleinfo.OptionType == u'敲入看涨自动敲出赎回':
            TempOptionType = u'敲入看涨自动敲出赎回新起息'
        else:
            TempOptionType = singleinfo.OptionType
        tempMonte = OptionDict[TempOptionType](S, t2, Para['Vol'][1],
                                               Para['Rf'][1], Para['Dividend'][1], 'Price',
                                               singleinfo) * multicontract
        tempGreeks = OptionDict[TempOptionType](S, t2, Para['Vol'][1],
                                                Para['Rf'][1], Para['Dividend'][1], 'All', singleinfo) * multicontract
        tempVanna = (OptionDict[TempOptionType](S, t2, Para['Vol'][1] + 0.001,
                                                Para['Rf'][1], Para['Dividend'][1], 'Delta', singleinfo) -
                     OptionDict[TempOptionType](S, t2, Para['Vol'][1] - 0.001,
                                                Para['Rf'][1], Para['Dividend'][1], 'Delta',
                                                singleinfo)) / 0.002 * multicontract
        tempZomma = (OptionDict[TempOptionType](S, t2, Para['Vol'][1] + 0.001,
                                                Para['Rf'][1], Para['Dividend'][1], 'Gamma', singleinfo) -
                     OptionDict[TempOptionType](S, t2, Para['Vol'][1] - 0.001,
                                                Para['Rf'][1], Para['Dividend'][1], 'Gamma',
                                                singleinfo)) / 0.002 * multicontract
        tempVolga = (OptionDict[TempOptionType](S, t2, Para['Vol'][1] + 0.001,
                                                Para['Rf'][1], Para['Dividend'][1], 'Vega', singleinfo) -
                     OptionDict[TempOptionType](S, t2, Para['Vol'][1] - 0.001,
                                                Para['Rf'][1], Para['Dividend'][1], 'Vega',
                                                singleinfo)) / 0.002 * multicontract
        return np.concatenate([[tempMonte], tempGreeks, [tempVanna, tempZomma, tempVolga]])

    def AccrualHist(self, aninfo, HistDate, HistType='All'):
        # 非区间累计结构或者未到区间累积时间为0
        if (HistDate < aninfo.BeginDate) or (u'累积' not in aninfo.OptionType):
            return 0
        NumTDays = int(of.mytdate(aninfo.BeginDate, aninfo.EndDate) * YearDays) + 1
        PriceSeri = np.array(w.wsd(aninfo.Code, 'close', aninfo.BeginDate, HistDate).Data[0])
        if aninfo.OptionType == u'区间累积':
            HistPrice = (sum((PriceSeri >= aninfo.InitialS * aninfo.LstrikeRatio)
                             & (PriceSeri <= aninfo.InitialS * aninfo.HstrikeRatio))
                         * aninfo.InitialS * aninfo.Rebate / NumTDays)
        elif aninfo.OptionType == u'三层累积看涨':
            HistPrice = (sum((PriceSeri >= aninfo.InitialS * aninfo.LstrikeRatio)
                             & (PriceSeri < aninfo.s0 * aninfo.HstrikeRatio))
                         * aninfo.InitialS * aninfo.Rebate +
                         sum(PriceSeri >= aninfo.InitialS * aninfo.HstrikeRatio)
                         * aninfo.InitialS * aninfo.Rebate2) / NumTDays
        elif aninfo.OptionType == u'四层累积看涨':
            HistPrice = (sum((PriceSeri >= aninfo.InitialS * aninfo.LstrikeRatio)
                             & (PriceSeri < aninfo.InitialS * aninfo.HstrikeRatio))
                         * aninfo.InitialS * aninfo.Rebate +
                         sum((PriceSeri >= aninfo.InitialS * aninfo.HstrikeRatio)
                             & (PriceSeri < aninfo.InitialS * aninfo.LtouchRatio))
                         * aninfo.InitialS * aninfo.Rebate2 +
                         sum(PriceSeri >= aninfo.InitialS * aninfo.LtouchRatio)
                         * aninfo.InitialS * aninfo.Rebate3) / NumTDays
        elif aninfo.OptionType == u'五层区间累积':
            HistPrice = (sum((PriceSeri >= aninfo.InitialS * aninfo.LstrikeRatio)
                             & (PriceSeri < aninfo.InitialS * aninfo.HstrikeRatio))
                         * aninfo.InitialS * aninfo.Rebate +
                         sum((PriceSeri >= aninfo.InitialS * aninfo.HstrikeRatio)
                             & (PriceSeri < aninfo.InitialS * aninfo.LtouchRatio))
                         * aninfo.InitialS * aninfo.Rebate2 +
                         sum((PriceSeri >= aninfo.InitialS * aninfo.LtouchRatio)
                             & (PriceSeri < aninfo.InitialS * aninfo.HtouchRatio))
                         * aninfo.InitialS * aninfo.Rebate3 +
                         sum(PriceSeri >= aninfo.InitialS * aninfo.HtouchRatio)
                         * aninfo.InitialS * aninfo.Rebate4) / NumTDays
        elif aninfo.OptionType == u'四层区间累积':
            HistPrice = (sum((PriceSeri >= aninfo.InitialS * aninfo.LstrikeRatio)
                             & (PriceSeri < aninfo.InitialS * aninfo.HstrikeRatio))
                         * aninfo.InitialS * aninfo.Rebate +
                         sum((PriceSeri >= aninfo.InitialS * aninfo.HstrikeRatio)
                             & (PriceSeri < aninfo.InitialS * aninfo.LtouchRatio))
                         * aninfo.InitialS * aninfo.Rebate2 +
                         sum((PriceSeri >= aninfo.InitialS * aninfo.LtouchRatio)
                             & (PriceSeri < aninfo.InitialS * aninfo.HtouchRatio))
                         * aninfo.InitialS * aninfo.Rebate3) / NumTDays
        else:
            raise ValueError('The OptionType May not Exist')

        if (HistType == 'MontePrice') or (HistType == 'Price'):
            return HistPrice
        elif HistType == 'All':
            return np.array([HistPrice, 0, 0, 0, 0])
        else:
            return 0

    def GetKnockIn(self):
        # 取 KnockInDate 所在列
        knockincol = np.nonzero(self.opinfo.columns == 'KnockInDate')[0][0] + 1
        wb = openpyxl.load_workbook(self.infoname)
        ws = wb['Sheet1']
        for i in range(np.size(self.opinfo, axis=0)):
            singleinfo = self.opinfo.iloc[i, :].copy()
            print("Start Get Knock In ", singleinfo.ContractId)
            if singleinfo.Code in self.PriceDict:
                S = self.PriceDict[singleinfo.Code][1]
            else:
                PreS = w.wsd(singleinfo.Code, 'close', self.TDate_yesterday, self.TDate_yesterday).Data[0][0]
                S = w.wsd(singleinfo.Code, 'close', self.TDate_today, self.TDate_today).Data[0][0]
                self.PriceDict[singleinfo.Code] = (PreS, S)
            if singleinfo.SettleMethod == 'TWAP' and singleinfo.ActualEndDate == self.TDate_today:
                S = singleinfo.TWAPSettlePrice
            IfKnockIn = 0
            if singleinfo.KnockInDate == 'NoKnockIn':
                if singleinfo.OptionType == u'敲入看涨自动敲出赎回':
                    IfKnockIn = (1 if S >= singleinfo.InitialS * singleinfo.HstrikeRatio else 0)
                elif singleinfo.OptionType == u'雪球看涨自动敲出赎回' or singleinfo.OptionType == u'逐步调整雪球看涨自动敲出赎回' or singleinfo.OptionType == u'收益凭证雪球':
                    """此处按收盘价小于敲入障碍价记为敲入触发，后续可以做修改"""
                    IfKnockIn = (1 if S < singleinfo.InitialS * singleinfo.LtouchRatio else 0)
            if IfKnockIn:
                ws.cell(row=singleinfo['ID'] + 1, column=knockincol, value=self.TDate_today)
                ws.cell(row=singleinfo['ID'] + 1, column=knockincol).number_format = 'mm-dd-yy'
                self.opinfo['KnockInDate'][self.opinfo['ID'] == singleinfo['ID']] = self.TDate_today
                self.allopinfo['KnockInDate'][self.allopinfo['ID'] == singleinfo['ID']] = self.TDate_today
            print("End Get Knock In ", singleinfo.ContractId)
        wb.save(self.infoname)

    def GetTouchExpire(self):
        # 取 ActualEndDate和ExerciseFee 所在列
        touchcol = np.nonzero(self.opinfo.columns == 'ActualEndDate')[0][0] + 1
        exercisecol = np.nonzero(self.opinfo.columns == 'ExerciseFee')[0][0] + 1
        wb = openpyxl.load_workbook(self.infoname)
        ws = wb['Sheet1']
        for i in range(np.size(self.opinfo, axis=0)):
            singleinfo = self.opinfo.iloc[i, :].copy()
            print("Start Get Knock Out ", singleinfo.ContractId)
            if singleinfo.Code in self.PriceDict:
                S = self.PriceDict[singleinfo.Code][1]
            else:
                PreS = w.wsd(singleinfo.Code, 'close', self.TDate_yesterday, self.TDate_yesterday).Data[0][0]
                S = w.wsd(singleinfo.Code, 'close', self.TDate_today, self.TDate_today).Data[0][0]
                self.PriceDict[singleinfo.Code] = (PreS, S)
            if singleinfo.SettleMethod == 'TWAP' and singleinfo.ActualEndDate == self.TDate_today:
                S = singleinfo.TWAPSettlePrice
            IfKnockOut = 0
            tempEndDate = singleinfo.ActualEndDate
            if (singleinfo.OptionType == u'单向鲨鱼鳍看跌') | (singleinfo.OptionType == u'美式二元看跌'):
                IfKnockOut = (1 if (S <= singleinfo.InitialS * singleinfo.LtouchRatio)
                                   & (tempEndDate > self.TDate_today) else 0)
            elif (singleinfo.OptionType == u'单向鲨鱼鳍看涨') | (singleinfo.OptionType == u'美式二元看涨')|(singleinfo.OptionType == u'美式二元向上不触碰'):
                IfKnockOut = (1 if (S >= singleinfo.InitialS * singleinfo.HtouchRatio)
                                   & (tempEndDate > self.TDate_today) else 0)
            elif (singleinfo.OptionType == u'美式二元看涨最高价'):
                # if singleinfo.Code=='000905.SH':
                #     # S = w.wsd(singleinfo.Code, 'high', self.TDate_today,self.TDate_today).Data[0][0]
                #     S = 4849.18
                # elif singleinfo.Code=='000300.SH':
                #     S = 3311.50
                # else:
                S = w.wsd(singleinfo.Code, 'high', self.TDate_today, self.TDate_today).Data[0][0]
                IfKnockOut = (1 if (S >= singleinfo.InitialS * singleinfo.HtouchRatio)
                                   & (tempEndDate > self.TDate_today) else 0)
            elif (singleinfo.OptionType == u'美式二元看跌最低价'):
                # if singleinfo.Code=='000905.SH':
                #     S = 4849.18
                # else:
                S = w.wsd(singleinfo.Code, 'low', self.TDate_today, self.TDate_today).Data[0][0]
                IfKnockOut = (1 if (S <= singleinfo.InitialS * singleinfo.LtouchRatio)
                                   & (tempEndDate > self.TDate_today) else 0)
            elif (singleinfo.OptionType == u'双向鲨鱼鳍') | (singleinfo.OptionType == u'美式双触碰') | (
                    singleinfo.OptionType == u'美式双不触碰'):
                IfKnockOut = (1 if ((S >= singleinfo.InitialS * singleinfo.HtouchRatio)
                                    | (S <= singleinfo.InitialS * singleinfo.LtouchRatio))
                                   & (tempEndDate > self.TDate_today) else 0)
            elif singleinfo.OptionType == u'不对称双向鲨鱼鳍':
                IfKnockOut = (1 if ((S >= singleinfo.InitialS * singleinfo.HtouchRatio)
                                    | (S <= singleinfo.InitialS * singleinfo.LtouchRatio))
                                   & (tempEndDate > self.TDate_today) else 0)
            elif singleinfo.OptionType == u'看涨自动敲出赎回':
                ObDate = sorted(singleinfo.ObSerDate.split('\n'))
                ObDate = list(map(lambda x: datetime.strptime(x, '%Y/%m/%d'), ObDate))
                IfKnockOut = (1 if ((S >= singleinfo.InitialS * singleinfo.HtouchRatio)
                                    & (self.TDate_today in ObDate)) else 0)
            elif singleinfo.OptionType == u'雪球看涨自动敲出赎回':
                ObDate = sorted(singleinfo.ObSerDate.split('\n'))
                ObDate = list(map(lambda x: datetime.strptime(x, '%Y/%m/%d'), ObDate))
                IfKnockOut = (1 if ((S >= singleinfo.InitialS * singleinfo.HtouchRatio)
                                    & (self.TDate_today in ObDate)) else 0)
            elif singleinfo.OptionType == u'逐步调整雪球看涨自动敲出赎回':
                # 障碍价在变化，需要判断当前日期对应的障碍价
                ObDate = sorted(singleinfo.ObSerDate.split('\n'))
                ObDate = list(map(lambda x: datetime.strptime(x, '%Y/%m/%d'), ObDate))
                if self.TDate_today in ObDate:
                    num = np.nonzero(np.array(ObDate) == self.TDate_today)[0][0]
                    KSeri = (singleinfo.HtouchRatio + num * singleinfo.Rebate3) * singleinfo.InitialS
                    IfKnockOut = 1 if S >= KSeri else 0
            elif singleinfo.OptionType == u'看跌自动敲出赎回':
                ObDate = sorted(singleinfo.ObSerDate.split('\n'))
                ObDate = list(map(lambda x: datetime.strptime(x, '%Y/%m/%d'), ObDate))
                IfKnockOut = (1 if ((S <= singleinfo.InitialS * singleinfo.LtouchRatio)
                                    & (self.TDate_today in ObDate)) else 0)
            elif singleinfo.OptionType == u'看涨自动敲出赎回转看涨':
                ObDate = sorted(singleinfo.ObSerDate.split('\n'))
                ObDate = list(map(lambda x: datetime.strptime(x, '%Y/%m/%d'), ObDate))
                IfKnockOut = (1 if ((S >= singleinfo.InitialS * singleinfo.HtouchRatio)
                                    & (self.TDate_today in ObDate)) else 0)
            elif singleinfo.OptionType == u'逐步调整看涨自动敲出赎回':
                ObDate = sorted(singleinfo.ObSerDate.split('\n'))
                ObDate = list(map(lambda x: datetime.strptime(x, '%Y/%m/%d'), ObDate))
                if self.TDate_today in ObDate:
                    num = np.nonzero(np.array(ObDate) == self.TDate_today)[0][0]
                    KSeri = (singleinfo.HtouchRatio + num * singleinfo.Rebate3) * singleinfo.InitialS
                    IfKnockOut = 1 if S >= KSeri else 0
            elif singleinfo.OptionType == u'敲入看涨自动敲出赎回':
                if singleinfo.KnockInDate == 'NoKnockIn':
                    IfKnockOut = 0
                else:
                    if singleinfo.KnockInDate <= self.TDate_today:
                        ObDate = sorted(singleinfo.ObSerDate.split('\n'))
                        ObDate = list(map(lambda x: datetime.strptime(x, '%Y/%m/%d'), ObDate))
                        IfKnockOut = (1 if ((S >= singleinfo.InitialS * singleinfo.HtouchRatio)
                                            & (self.TDate_today in ObDate)) else 0)
            if IfKnockOut:
                tempEndDate = self.TDate_today
                ws.cell(row=singleinfo['ID'] + 1, column=touchcol, value=self.TDate_today)
                ws.cell(row=singleinfo['ID'] + 1, column=touchcol).number_format = 'mm-dd-yy'
                self.allopinfo['ActualEndDate'][self.allopinfo['ID'] == singleinfo['ID']] = self.TDate_today
            if tempEndDate <= self.TDate_today:
                if singleinfo.OptionType == u'雪球看涨自动敲出赎回' or singleinfo.OptionType == u'逐步调整雪球看涨自动敲出赎回':
                    SettleChazhi = (DealTradeDate.timedate(singleinfo.DealDate) - DealTradeDate.timedate(singleinfo.EndDate)).days
                    # singleinfo.Premium为预付金利息，负的话是我们付客户,三部分收益，1、敲出损益，2、敲入损益，
                    exercisefee = (singleinfo.Notamt * singleinfo.Pratio * (singleinfo.Rebate - singleinfo.Premium*singleinfo.Pos - singleinfo.PrePayFee*singleinfo.PrePayFeeRevenue) * singleinfo.Pos
                                   * ((self.TDate_today - DealTradeDate.timedate(
                                singleinfo.BeginDate)).days + SettleChazhi) / singleinfo.YearDays) * IfKnockOut+\
                                  singleinfo.Notamt*(self.PriceDict[singleinfo.Code][1] - singleinfo.HstrikeRatio)/singleinfo.InitialS*singleinfo.Pos*(1-IfKnockOut)
                elif singleinfo.OptionType == '收益凭证雪球':
                    # 目前行权费中包含销售奖励，后续可剥离
                    SettleChazhi = (DealTradeDate.timedate(singleinfo.DealDate) - DealTradeDate.timedate(
                        singleinfo.EndDate)).days
                    SalesFee = singleinfo.Notamt * singleinfo.Pratio * singleinfo.SalesFee
                    exercisefee = (singleinfo.Notamt * singleinfo.Pratio * (
                                singleinfo.Rebate - singleinfo.Premium*singleinfo.Pos - singleinfo.PrePayFee * singleinfo.PrePayFeeRevenue) * singleinfo.Pos
                                   * ((self.TDate_today - DealTradeDate.timedate(
                                singleinfo.BeginDate)).days + SettleChazhi) / singleinfo.YearDays) * IfKnockOut + \
                                  singleinfo.Notamt * (self.PriceDict[singleinfo.Code][
                                                           1] - singleinfo.HstrikeRatio) / singleinfo.InitialS * singleinfo.Pos * (
                                              1 - IfKnockOut) + SalesFee

                elif singleinfo.OptionType == u'看涨自动敲出赎回':
                    SettleChazhi = (DealTradeDate.timedate(singleinfo.DealDate) - DealTradeDate.timedate(singleinfo.EndDate)).days
                    # ReturnFee为期权费利息
                    ReturnFee = (singleinfo.Notamt / singleinfo.YearDays * singleinfo.Maturity * singleinfo.Premium
                                 * singleinfo.Rebate2 * ((self.TDate_today - DealTradeDate.timedate(
                                singleinfo.BeginDate)).days + SettleChazhi) / singleinfo.YearDays)
                    exercisefee = (singleinfo.Notamt * singleinfo.Pratio * singleinfo.Rebate * singleinfo.Pos
                                   * ((self.TDate_today - DealTradeDate.timedate(
                                singleinfo.BeginDate)).days + SettleChazhi) / singleinfo.YearDays + ReturnFee) * IfKnockOut
                    # 按道理参与率已经算进期权费中，所以此出不应该乘以参与率

                elif singleinfo.OptionType == u'看跌自动敲出赎回':
                    SettleChazhi = (DealTradeDate.timedate(singleinfo.DealDate) - DealTradeDate.timedate(singleinfo.EndDate)).days
                    ReturnFee = (singleinfo.Notamt / singleinfo.YearDays * singleinfo.Maturity * singleinfo.Premium
                                 * singleinfo.Rebate2 * ((self.TDate_today - DealTradeDate.timedate(
                                singleinfo.BeginDate)).days + SettleChazhi) / singleinfo.YearDays)
                    exercisefee = (singleinfo.Notamt * singleinfo.Pratio * singleinfo.Rebate * singleinfo.Pos
                                   * ((self.TDate_today - DealTradeDate.timedate(
                                singleinfo.BeginDate)).days + SettleChazhi) / singleinfo.YearDays + ReturnFee) * IfKnockOut
                    # 按道理参与率已经算进期权费中，所以此出不应该乘以参与率
                elif singleinfo.OptionType == u'看涨自动敲出赎回转看涨':
                    SettleChazhi = (DealTradeDate.timedate(singleinfo.DealDate) - DealTradeDate.timedate(singleinfo.EndDate)).days
                    if IfKnockOut:
                        ReturnFee = (singleinfo.Notamt / singleinfo.YearDays * singleinfo.Maturity * singleinfo.Premium
                                     * singleinfo.Rebate2 * ((self.TDate_today - DealTradeDate.timedate(
                                    singleinfo.BeginDate)).days + SettleChazhi) / singleinfo.YearDays)
                        exercisefee = (singleinfo.Notamt * singleinfo.Pratio * singleinfo.Rebate * singleinfo.Pos
                                       * ((self.TDate_today - DealTradeDate.timedate(
                                    singleinfo.BeginDate)).days + SettleChazhi) / singleinfo.YearDays + ReturnFee)
                    else:
                        multicontract = (singleinfo.Notamt * singleinfo.Maturity /
                                         singleinfo.YearDays * singleinfo.Pratio * singleinfo.Pos
                                         if singleinfo.IfActual == 0 else
                                         singleinfo.Notamt * singleinfo.Pratio * singleinfo.Pos) / singleinfo.InitialS
                        exercisefee = (OptionDict[singleinfo.OptionType](S, 0, 0.3, 0, 0, 'Price', singleinfo) +
                                       self.AccrualHist(singleinfo, self.TDate_yesterday, 'Price')) * multicontract
                    # 按道理参与率已经算进期权费中，所以此出不应该乘以参与率
                elif singleinfo.OptionType == u'逐步调整看涨自动敲出赎回':
                    SettleChazhi = (DealTradeDate.timedate(singleinfo.DealDate) - DealTradeDate.timedate(singleinfo.EndDate)).days
                    ReturnFee = (singleinfo.Notamt / singleinfo.YearDays * singleinfo.Maturity * singleinfo.Premium
                                 * singleinfo.Rebate2 * ((self.TDate_today - DealTradeDate.timedate(
                                singleinfo.BeginDate)).days + SettleChazhi) / singleinfo.YearDays)
                    exercisefee = (singleinfo.Notamt * singleinfo.Pratio * singleinfo.Rebate * singleinfo.Pos
                                   * ((self.TDate_today - DealTradeDate.timedate(
                                singleinfo.BeginDate)).days + SettleChazhi) / singleinfo.YearDays + ReturnFee) * IfKnockOut
                else:
                    multicontract = (singleinfo.Notamt * singleinfo.Maturity /
                                     singleinfo.YearDays * singleinfo.Pratio * singleinfo.Pos
                                     if singleinfo.IfActual == 0 else
                                     singleinfo.Notamt * singleinfo.Pratio * singleinfo.Pos) / singleinfo.InitialS
                    exercisefee = (OptionDict[singleinfo.OptionType](S, 0, 0.3, 0, 0, 'Price', singleinfo) +
                                   self.AccrualHist(singleinfo, self.TDate_yesterday, 'Price')) * multicontract
                ws.cell(row=singleinfo['ID'] + 1, column=exercisecol, value=exercisefee)
                self.allopinfo['ExerciseFee'][self.allopinfo['ID'] == singleinfo['ID']] = exercisefee
        wb.save(self.infoname)

    def GatherExoticReport(self):
        gatherfile = '.\ProfitLossAnalyst\OTC_PL_Decompose_Summary.xlsx'
        temppl = [self.TDate_today,
                  self.PoolInfo['TotalP&L'][self.PoolInfo['Date'] == 'BiasP&L'].iloc[0],
                  self.PoolInfo['TotalP&L'][self.PoolInfo['Date'] == 'PositionP&L'].iloc[0],
                  self.PoolInfo['TotalP&L'][self.PoolInfo['Date'] == 'DeltaP&L'].iloc[0],
                  self.PoolInfo['TotalP&L'][self.PoolInfo['Date'] == 'GammaP&L'].iloc[0],
                  self.PoolInfo['TotalP&L'][self.PoolInfo['Date'] == 'ThetaP&L'].iloc[0],
                  self.PoolInfo['TotalP&L'][self.PoolInfo['Date'] == 'Gamma+ThetaP&L'].iloc[0],
                  self.PoolInfo['TotalP&L'][self.PoolInfo['Date'] == 'VegaP&L'].iloc[0],
                  self.PoolInfo['TotalP&L'][self.PoolInfo['Date'] == 'GreeksPL'].iloc[0],
                  self.PoolInfo['SSigB'][self.PoolInfo['Date'] == 'GreeksError'].iloc[0],
                  self.PoolInfo['TotalExistP&L'][self.PoolInfo['Date'] == 'Profit&Loss'].iloc[0],
                  self.PoolInfo['NewP&L'][self.PoolInfo['Date'] == self.TDate_yesterday].iloc[0],
                  self.PoolInfo['NewP&L'][self.PoolInfo['Date'] == self.TDate_today].iloc[0],
                  self.PoolInfo['NewP&L'][self.PoolInfo['Date'] == 'Profit&Loss'].iloc[0],
                  self.PoolInfo['TotalP&L'][self.PoolInfo['Date'] == 'Profit&Loss'].iloc[0]]
        wb = openpyxl.load_workbook(gatherfile)
        num_form = '_ * #,##0_ ;_ * \\-#,##0_ ;_ * "-"??_ ;_ @_ '
        try:
            ws = wb[self.UNCode]
            print("start make Old %s PL file"%self.UNCode)
            ifend = 1
            for i in range(1, len(list(ws.rows))):
                if ws.cell(row=i + 1, column=1).value >= self.TDate_today:
                    ifend = 0
                    break
            sindex = i + ifend + 1
            for tindex, tcell in enumerate(temppl):
                ws.cell(row=sindex, column=tindex + 1, value=tcell)
                ws.cell(row=sindex, column=tindex + 1).number_format = num_form
                if isinstance(tcell, datetime):
                    ws.cell(row=sindex, column=tindex + 1).number_format = 'mm-dd-yy'
                sumvalue = temppl[-1] + ws.cell(row=sindex - 1, column=len(temppl) + 1).value
                ws.cell(row=sindex, column=len(temppl) + 1, value=sumvalue)
                ws.cell(row=sindex, column=len(temppl) + 1).number_format = num_form
                wb.save(gatherfile)
            print("%s Old PL file made done" % self.UNCode)
        except:
            ws = wb.create_sheet(self.UNCode)
            print("start make New %s PL file" % self.UNCode)
            ws.cell(row=1, column=1).value = 'TDate'
            ws.cell(row=1, column=2).value = 'BiasPL'
            ws.cell(row=1, column=3).value = 'PositionPL'
            ws.cell(row=1, column=4).value = 'DeltaPL'
            ws.cell(row=1, column=5).value = 'GammaPL'
            ws.cell(row=1, column=6).value = 'ThetaPL'
            ws.cell(row=1, column=7).value = 'GammaThetaPL'
            ws.cell(row=1, column=8).value = 'VegaPL'
            ws.cell(row=1, column=9).value = 'GreeksPL'
            ws.cell(row=1, column=10).value = 'GreeksError'
            ws.cell(row=1, column=11).value = 'TotalExistPL'
            ws.cell(row=1, column=12).value = 'NewFee'
            ws.cell(row=1, column=13).value = 'NewValue'
            ws.cell(row=1, column=14).value = 'NewPL'
            ws.cell(row=1, column=15).value = 'TotalPL'
            ws.cell(row=1, column=16).value = 'SumPL'
            sindex = 2
            for tindex, tcell in enumerate(temppl):
                ws.cell(row=sindex, column=tindex + 1, value=tcell)
                ws.cell(row=sindex, column=tindex + 1).number_format = num_form
                if isinstance(tcell, datetime):
                    ws.cell(row=sindex, column=tindex + 1).number_format = 'mm-dd-yy'
                sumvalue = temppl[-1]
                try:
                    ws.cell(row=sindex, column=len(temppl) + 1, value=sumvalue)
                    ws.cell(row=sindex, column=len(temppl) + 1).number_format = num_form
                except:
                    print(self.UNCode, sumvalue, sindex, len(temppl) + 1)
                wb.save(gatherfile)
            print("%s New PL file made done" % self.UNCode)


def CalAveCorr(UnCode, HVNum, TDate):
    # 这里应该是计算相关系数矩阵，或者平均相关系数，但此处先计算单一相关系数
    for j, tempCode in enumerate(UnCode):
        if j == 0:
            codestr = tempCode
        else:
            codestr = codestr + ',' + tempCode
    BeginDate = w.tdaysoffset(-HVNum - 1, TDate, "").Data[0][0]
    Ps = np.array(w.wsd(codestr, 'close', BeginDate, TDate).Data)
    LogReturn = np.log(Ps)[:, 1:] - np.log(Ps)[:, :-1]
    return np.round(np.corrcoef(LogReturn[:, :-1])[0, 1], 4), np.round(np.corrcoef(LogReturn[:, 1:])[0, 1], 4)


class MultiUnderlyingExoticReport():
    def __init__(self, underlyingcode, Maturity, NotBias=False):
        self.NotBias = NotBias
        self.UNCode = underlyingcode
        self.TDate_today = DealTradeDate.timedate(w.tdaysoffset(0, Maturity).Data[0][0])
        self.TDate_yesterday = DealTradeDate.timedate(w.tdaysoffset(-1, self.TDate_today).Data[0][0])
        # self.TDate_before_yesterday = DealTradeDate.timedate(w.tdaysoffset(-1,self.TDate_yesterday).Data[0][0])
        self.filename = (os.getcwd() + '\\ProfitLossAnalyst\\' +
                         self.TDate_today.strftime('%Y%m%d') + self.UNCode + 'PL.xlsx')

        self.infoname = ('.\Contracts_and_trade\MyTotalContracts.xlsx')
        # self.infoname = ('C:\\Users\\yanghua\\Desktop\\MyTotalContracts.xlsx')
        try:
            totalopinfo = pd.read_excel(self.infoname)
        except:
            raise ValueError('Please check the contract excel whether exist!')

        self.allopinfo = totalopinfo[totalopinfo.Underlying == self.UNCode]
        self.opinfo = self.allopinfo[((self.allopinfo.ActualEndDate >= self.TDate_today)
                                      & (totalopinfo.BeginDate < self.TDate_today))]
        self.newinfo = self.allopinfo[self.allopinfo.BeginDate == self.TDate_today]
        # 这里输入Code的昨日和今日的价格，格式'code':(昨日价格：今日价格)
        self.PriceDict = pd.DataFrame(index=[self.TDate_yesterday, self.TDate_today])
        self.FutureDict = pd.DataFrame(index=[self.TDate_yesterday, self.TDate_today])
        self.UnderNum = len(CorrProduct[self.UNCode])
        self.PartialCode = CorrProduct[self.UNCode][:, 0]
        self.CorrS0 = pd.DataFrame(index=self.PartialCode)
        self.PartialUnder = CorrProduct[self.UNCode][:, 1]

        for i in range(self.UnderNum):
            tempS_yesterday = w.wsd(self.PartialCode[i], 'close', self.TDate_yesterday, self.TDate_yesterday).Data[0][0]
            if CorrProduct[self.UNCode][i, 0] == 'AU9999':  # 这里需要手输AU9999收盘价
                tempS_today = (AU9999P if self.TDate_today >= DealTradeDate.timedate(datetime.today()) else
                               w.wsd(self.PartialCode[i], 'close', self.TDate_today, self.TDate_today).Data[0][0])
            else:
                tempS_today = w.wsd(self.PartialCode[i], 'close', self.TDate_today, self.TDate_today).Data[0][0]
            self.PriceDict[self.PartialCode[i]] = (tempS_yesterday, tempS_today)

            tempS_yesterday = \
            w.wsd(FutureCode[self.PartialUnder[i]], 'close', self.TDate_yesterday, self.TDate_yesterday).Data[0][0]
            tempS_today = w.wsd(FutureCode[self.PartialUnder[i]], 'close', self.TDate_today, self.TDate_today).Data[0][
                0]
            self.FutureDict[FutureCode[self.PartialUnder[i]]] = (tempS_yesterday, tempS_today)

        tempcorrinfo = self.allopinfo[self.allopinfo.ActualEndDate >= self.TDate_today]
        for tempcorrindex in range(np.size(tempcorrinfo, axis=0)):
            tempcorr = tempcorrinfo.iloc[tempcorrindex, :]
            tempseri = [w.wsd(self.PartialCode[i], 'close', tempcorr.BeginDate, tempcorr.BeginDate).Data[0][0]
                        for i in range(self.UnderNum)]
            self.CorrS0[tempcorr.Code] = tempseri
            # 读取参数表里面的信息
        # TODO 由于多个标的的期权没有存续的，所以这里的读取参数暂不修改，后续如果有多个标的的期权的话需要进行修改
        paraexcel_yesterday = ('.\每日报价\%s每日报价.xlsx' % self.TDate_yesterday.strftime('%Y%m%d'))
        paraexcel_today = ('.\每日报价\%s每日报价.xlsx' % self.TDate_today.strftime('%Y%m%d'))
        # paraexcel_yesterday = (u'C:\\Users\\yanghua\\Desktop\\每日报价\\'+
        #                   self.TDate_yesterday.strftime('%Y%m%d')+u'每日报价.xlsx')
        # paraexcel_today = (u'C:\\Users\\yanghua\\Desktop\\每日报价\\'+
        #                  self.TDate_today.strftime('%Y%m%d')+u'每日报价.xlsx')
        para_yesterday = pd.read_excel(paraexcel_yesterday, skiprows=3, usecols='M:W', index_col=0)
        para_today = pd.read_excel(paraexcel_today, skiprows=3, usecols='M:W', index_col=0)

        self.RiskPara1 = pd.DataFrame(index=['Vol', 'Rf', 'Dividend', 'Rho'], columns=self.PartialUnder)
        self.RiskPara2 = pd.DataFrame(index=['Vol', 'Rf', 'Dividend', 'Rho'], columns=self.PartialUnder)
        self.TradePara1 = pd.DataFrame(index=['Vol', 'Rf', 'Dividend', 'Rho'], columns=self.PartialUnder)
        self.TradePara2 = pd.DataFrame(index=['Vol', 'Rf', 'Dividend', 'Rho'], columns=self.PartialUnder)
        for i in range(self.UnderNum):
            paracode = ParaTable[self.PartialUnder[i]] if self.PartialUnder[i] in ParaTable else self.PartialUnder[i]
            self.RiskPara1[self.PartialUnder[i]]['Vol'] = para_yesterday['Vol'][paracode] / 100
            self.RiskPara2[self.PartialUnder[i]]['Vol'] = para_today['Vol'][paracode] / 100
            self.RiskPara1[self.PartialUnder[i]]['Rf'] = para_yesterday['Rf'][paracode] / 100
            self.RiskPara2[self.PartialUnder[i]]['Rf'] = para_today['Rf'][paracode] / 100
            self.RiskPara1[self.PartialUnder[i]]['Dividend'] = para_yesterday['Dividend'][paracode] / 100
            self.RiskPara2[self.PartialUnder[i]]['Dividend'] = para_today['Dividend'][paracode] / 100

            self.TradePara1[self.PartialUnder[i]]['Vol'] = para_yesterday['TradeVol'][self.UNCode] / 100
            self.TradePara2[self.PartialUnder[i]]['Vol'] = para_yesterday['TradeVol'][self.UNCode] / 100
            self.TradePara1[self.PartialUnder[i]]['Dividend'] = 0
            self.TradePara2[self.PartialUnder[i]]['Dividend'] = 0
            self.TradePara1[self.PartialUnder[i]]['Rf'] = para_yesterday['Rf'][paracode] / 100 if para_yesterday['Rf'][
                                                                                                      paracode] != DepositRate else 0
            self.TradePara2[self.PartialUnder[i]]['Rf'] = para_today['Rf'][paracode] / 100 if para_today['Rf'][
                                                                                                  paracode] != DepositRate else 0

        self.RiskPara1[self.PartialUnder[0]]['Rho'], self.RiskPara2[self.PartialUnder[0]]['Rho'] = CalAveCorr(
            self.PartialCode, 60, self.TDate_today)
        # 目前先粗糙处理交易波动率，交易波动率方面就把波动率考虑成一个整体
        # 如果后续需要精细化处理，在此处进行修改

        self.RiskTableTitle = np.array(CorrRiskTableTitle)
        self.TradeTableTitle = np.array(CorrTradeTableTitle)
        DeltaStr = ['Delta' + str(i + 1) for i in range(self.UnderNum)]
        tempindex = np.nonzero(self.RiskTableTitle == 'RiskDeltaPL')[0][0]
        self.RiskTableTitle = np.insert(self.RiskTableTitle, tempindex, DeltaStr)
        self.TradeTableTitle = np.insert(self.TradeTableTitle, tempindex, DeltaStr)
        GammaStr = ['Gamma' + str(i + 1) for i in range(self.UnderNum)]
        tempindex = np.nonzero(self.RiskTableTitle == 'RiskGammaPL')[0][0]
        self.RiskTableTitle = np.insert(self.RiskTableTitle, tempindex, GammaStr)
        self.TradeTableTitle = np.insert(self.TradeTableTitle, tempindex, GammaStr)

        self.NewTableTitle = np.array(CorrNewTableTitle)
        RiskNewStr = ['RiskDelta' + str(i + 1) for i in range(self.UnderNum)]
        RiskNewStr.extend(['RiskGamma' + str(i + 1) for i in range(self.UnderNum)])
        tempindex = np.nonzero(self.NewTableTitle == 'RiskTheta')[0][0]
        self.NewTableTitle = np.insert(self.NewTableTitle, tempindex, RiskNewStr)
        TradeNewStr = ['TradeDelta' + str(i + 1) for i in range(self.UnderNum)]
        TradeNewStr.extend(['TradeGamma' + str(i + 1) for i in range(self.UnderNum)])
        tempindex = np.nonzero(self.NewTableTitle == 'TradeTheta')[0][0]
        self.NewTableTitle = np.insert(self.NewTableTitle, tempindex, TradeNewStr)

    def genreport(self):
        writer = pd.ExcelWriter(self.filename, engine='xlsxwriter', datetime_format='yyyy/mm/dd')
        workbook = writer.book
        format0 = workbook.add_format({'font_name': 'Arial Unicode MS',
                                       'num_format': '#,##0.00', 'font_size': 10})
        format1 = workbook.add_format({'font_name': 'Arial Unicode MS',
                                       'num_format': '0.00%', 'font_size': 10})
        format2 = workbook.add_format({'font_name': 'Arial Unicode MS',
                                       'num_format': '#,##0', 'font_size': 10})
        format3 = workbook.add_format({'bold': True, 'font_name': 'Arial Unicode MS',
                                       'num_format': '#,##0', 'font_size': 10,
                                       'bg_color': 'yellow'})
        format4 = workbook.add_format({'bold': True, 'font_name': 'Arial Unicode MS',
                                       'num_format': '_(* #,##0_);_(* (#,##0);_(* "-"_);_(@_)',
                                       'font_size': 10, 'bg_color': 'yellow'})

        if self.opinfo.shape[0] != 0:
            riskgreeks = np.round(self.calgreeks(), 2)
            tradegreeks = np.round(self.calgreeks('Trade'), 2)
            riskgreeks.to_excel(writer, sheet_name=u'风控参数', startcol=4, index=False)
            RiskParaSheet = writer.sheets[u'风控参数']
            RiskParaSheet.set_column('A:AZ', 14, format0)
            tradegreeks.to_excel(writer, sheet_name=u'交易参数', startcol=4, index=False)
            TradeParaSheet = writer.sheets[u'交易参数']
            TradeParaSheet.set_column('A:AZ', 14, format0)
            self.GetTouchExpire()
        else:
            riskgreeks = pd.DataFrame(pd.Series({'contractid': 'sum', 'code': ''
                                                 }).reindex(index=self.RiskTableTitle, fill_value=0)).T
            tradegreeks = pd.DataFrame(pd.Series({'contractid': 'sum', 'code': ''
                                                  }).reindex(index=self.TradeTableTitle, fill_value=0)).T

        if self.newinfo.shape[0] != 0:
            newgreeks = self.newcalgreeks()
            newgreeks.to_excel(writer, sheet_name=u'新起息产品', startcol=4, index=False)
            NewContract = writer.sheets[u'新起息产品']
            NewContract.set_column('A:AZ', 14, format0)
        else:
            newgreeks = pd.DataFrame(pd.Series({'contractid': 'sum', 'code': ''
                                                }).reindex(index=self.NewTableTitle, fill_value=0)).T

        ParaDf = pd.DataFrame(index=[self.TDate_yesterday, self.TDate_today])
        for i in range(self.UnderNum):
            ParaDf[self.PartialUnder[i] + u'风控波动率'] = (self.RiskPara1[self.PartialUnder[i]]['Vol'],
                                                       self.RiskPara2[self.PartialUnder[i]]['Vol'])
            ParaDf[self.PartialUnder[i] + u'风控无风险利率'] = (self.RiskPara1[self.PartialUnder[i]]['Rf'],
                                                         self.RiskPara2[self.PartialUnder[i]]['Rf'])
            ParaDf[self.PartialUnder[i] + u'风控分红率'] = (self.RiskPara1[self.PartialUnder[i]]['Dividend'],
                                                       self.RiskPara2[self.PartialUnder[i]]['Dividend'])
            ParaDf[self.PartialUnder[i] + u'交易波动率'] = (self.TradePara1[self.PartialUnder[i]]['Vol'],
                                                       self.TradePara2[self.PartialUnder[i]]['Vol'])
            ParaDf[self.PartialUnder[i] + u'交易无风险利率'] = (self.TradePara1[self.PartialUnder[i]]['Rf'],
                                                         self.TradePara2[self.PartialUnder[i]]['Rf'])
            ParaDf[self.PartialUnder[i] + u'交易分红率'] = (self.TradePara1[self.PartialUnder[i]]['Dividend'],
                                                       self.TradePara2[self.PartialUnder[i]]['Dividend'])
        ParaDf[u'风控相关系数'] = (self.RiskPara1[self.PartialUnder[0]]['Rho'],
                             self.RiskPara2[self.PartialUnder[0]]['Rho'])
        if ~np.isnan(self.TradePara2.loc['Rho'][0]):
            ParaDf[u' 交易相关系数'] = (self.TradePara1[self.PartialUnder[0]]['Rho'],
                                  self.TradePara2[self.PartialUnder[0]]['Rho'])

        ParaDf = ParaDf.T
        BasicInfo = ParaDf.append(self.PriceDict.T)
        BasicInfo.to_excel(writer, sheet_name='Sheet1', startrow=1, index=True)

        self.allopinfo.to_excel(writer, sheet_name='Sheet1', startcol=4, index=False)
        worksheet1 = writer.sheets['Sheet1']
        worksheet1.set_column('A:AZ', 14)
        worksheet1.write_string(0, 1, u'日期1')
        worksheet1.write_string(0, 2, u'日期2')
        parawriteframe = 'B3:C' + str(len(ParaDf) + 2)
        worksheet1.conditional_format(parawriteframe, {'type': 'cell', 'criteria': '<', 'value': 5, 'format': format1})

        if (self.newinfo.shape[0] != 0) & (self.opinfo.shape[0] != 0):
            RiskParaSheet.conditional_format(parawriteframe,
                                             {'type': 'cell', 'criteria': '<', 'value': 5, 'format': format1})
            BasicInfo.to_excel(writer, sheet_name=u'风控参数', startrow=1, index=True)
            TradeParaSheet.conditional_format(parawriteframe,
                                              {'type': 'cell', 'criteria': '<', 'value': 5, 'format': format1})
            BasicInfo.to_excel(writer, sheet_name=u'交易参数', startrow=1, index=True)
            NewContract.conditional_format(parawriteframe,
                                           {'type': 'cell', 'criteria': '<', 'value': 5, 'format': format1})
            BasicInfo.to_excel(writer, sheet_name=u'新起息产品', startrow=1, index=True)
        elif (self.newinfo.shape[0] != 0) & (self.opinfo.shape[0] == 0):
            NewContract.conditional_format(parawriteframe,
                                           {'type': 'cell', 'criteria': '<', 'value': 5, 'format': format1})
            BasicInfo.to_excel(writer, sheet_name=u'新起息产品', startrow=1, index=True)
        elif (self.newinfo.shape[0] == 0) & (self.opinfo.shape[0] != 0):
            RiskParaSheet.conditional_format(parawriteframe,
                                             {'type': 'cell', 'criteria': '<', 'value': 5, 'format': format1})
            BasicInfo.to_excel(writer, sheet_name=u'风控参数', startrow=1, index=True)
            TradeParaSheet.conditional_format(parawriteframe,
                                              {'type': 'cell', 'criteria': '<', 'value': 5, 'format': format1})
            BasicInfo.to_excel(writer, sheet_name=u'交易参数', startrow=1, index=True)

        PoolTitle = ['RiskMonte', 'SSigB', 'SSigT', 'Risk-Trade', 'FutureP&L',
                     'TotalP&L', 'NewP&L', 'TotalExistP&L']
        # TODO change Excel to mysql
        ProfitLoss_Df = pd.read_excel('.\Contracts_and_trade\对冲头寸损益.xlsx',
                                      sheet_name=FuturePLCode[self.UNCode], index_col=0)
        # ProfitLoss_Df=pd.read_excel(u'C:\\Users\\yanghua\\Desktop\\对冲头寸损益.xlsx',
        #                   sheet_name=FuturePLCode[self.UNCode],index_col=0)
        FuturePL = ProfitLoss_Df[u'收盘价当日盈亏'][self.TDate_today]
        FutureDeltaPL = ProfitLoss_Df[u'昨日Delta盈亏'][self.TDate_today]
        FutureGammaPL = FuturePL - FutureDeltaPL
        self.PoolInfo = pd.DataFrame()
        TempInfoSeri = np.array([self.TDate_yesterday, self.TDate_today, 'Profit&Loss',
                                 'BiasP&L', 'PositionP&L', 'DeltaP&L',
                                 'GammaP&L', 'Theta', 'ThetaP&L', 'Gamma+ThetaP&L',
                                 'Vega', 'VegaP&L', 'GreeksPL', 'GreeksError'])
        DeltaStr = ['Delta' + str(i + 1) for i in range(self.UnderNum)]
        tempindex = np.nonzero(TempInfoSeri == 'BiasP&L')[0][0]
        TempInfoSeri = np.insert(TempInfoSeri, tempindex, DeltaStr)
        GammaStr = ['Gamma' + str(i + 1) for i in range(self.UnderNum)]
        tempindex = np.nonzero(TempInfoSeri == 'GammaP&L')[0][0]
        TempInfoSeri = np.insert(TempInfoSeri, tempindex, GammaStr)

        self.PoolInfo = pd.DataFrame(index=TempInfoSeri)
        PoolInfoLen = len(TempInfoSeri)

        tempfuturepl = pd.Series(index=TempInfoSeri)
        tempfuturepl['Profit&Loss'] = FuturePL
        tempfuturepl['BiasP&L'] = 0
        tempfuturepl['PositionP&L'] = 0
        tempfuturepl['DeltaP&L'] = FutureDeltaPL
        tempfuturepl['GammaP&L'] = FutureGammaPL
        tempfuturepl['ThetaP&L'] = 0
        tempfuturepl['Gamma+ThetaP&L'] = FutureGammaPL
        tempfuturepl['VegaP&L'] = 0
        tempfuturepl['GreeksPL'] = FuturePL
        tempfuturepl['GreeksError'] = 0
        self.PoolInfo['FutureP&L'] = tempfuturepl

        tempRiskMonte = np.append(np.array(
            [riskgreeks['RiskMonteValuePre'].iloc[-1],
             riskgreeks['RiskMonteValueCur'].iloc[-1] +
             newgreeks['RiskMonteValue'].iloc[-1],
             riskgreeks['RiskMontePL'].iloc[-1] +
             newgreeks['OptionFee'].iloc[-1] -
             newgreeks['RiskMonteValue'].iloc[-1]]),
            np.nan * np.ones(PoolInfoLen - 3))
        self.PoolInfo['RiskMonte'] = tempRiskMonte

        tempRiskCol = pd.Series(index=TempInfoSeri)
        tempRiskCol[self.TDate_yesterday] = riskgreeks['RiskPricePre'].iloc[-1]
        tempRiskCol[self.TDate_today] = riskgreeks['RiskPriceCur'].iloc[-1] + newgreeks['RiskPrice'].iloc[-1]
        tempRiskCol['Profit&Loss'] = (riskgreeks['RiskPL'].iloc[-1] +
                                      newgreeks['OptionFee'].iloc[-1] +
                                      newgreeks['RiskPrice'].iloc[-1])
        tempRiskCol['BiasP&L'] = riskgreeks['RiskBiasPL'].iloc[-1]
        tempRiskCol['PositionP&L'] = 0
        tempRiskCol['DeltaP&L'] = riskgreeks['RiskDeltaPL'].iloc[-1]
        tempRiskCol['GammaP&L'] = riskgreeks['RiskGammaPL'].iloc[-1]
        tempRiskCol['Theta'] = riskgreeks['RiskTheta'].iloc[-1] + newgreeks['RiskTheta'].iloc[-1]
        tempRiskCol['ThetaP&L'] = riskgreeks['RiskThetaPL'].iloc[-1]
        tempRiskCol['Gamma+ThetaP&L'] = riskgreeks['RiskGammaPL'].iloc[-1] + riskgreeks['RiskThetaPL'].iloc[-1]
        tempRiskCol['Vega'] = riskgreeks['RiskVega'].iloc[-1] + newgreeks['RiskVega'].iloc[-1]
        tempRiskCol['VegaP&L'] = riskgreeks['RiskVegaPL'].iloc[-1]
        tempRiskCol['GreeksPL'] = riskgreeks['RiskGreeksPL'].iloc[-1]
        tempRiskCol['GreeksError'] = riskgreeks['RiskPL'].iloc[-1] - riskgreeks['RiskGreeksPL'].iloc[-1]

        tempTradeCol = pd.Series(index=TempInfoSeri)
        tempTradeCol[self.TDate_yesterday] = tradegreeks['TradePricePre'].iloc[-1]
        tempTradeCol[self.TDate_today] = tradegreeks['TradePriceCur'].iloc[-1] + newgreeks['TradePrice'].iloc[-1]
        tempTradeCol['Profit&Loss'] = (tradegreeks['TradePL'].iloc[-1] +
                                       newgreeks['OptionFee'].iloc[-1] +
                                       newgreeks['TradePrice'].iloc[-1])
        tempTradeCol['BiasP&L'] = tradegreeks['TradeBiasPL'].iloc[-1]
        tempTradeCol['PositionP&L'] = 0
        tempTradeCol['DeltaP&L'] = tradegreeks['TradeDeltaPL'].iloc[-1]
        tempTradeCol['GammaP&L'] = tradegreeks['TradeGammaPL'].iloc[-1]
        tempTradeCol['Theta'] = tradegreeks['TradeTheta'].iloc[-1] + newgreeks['TradeTheta'].iloc[-1]
        tempTradeCol['ThetaP&L'] = tradegreeks['TradeThetaPL'].iloc[-1]
        tempTradeCol['Gamma+ThetaP&L'] = tradegreeks['TradeGammaPL'].iloc[-1] + tradegreeks['TradeThetaPL'].iloc[-1]
        tempTradeCol['Vega'] = tradegreeks['TradeVega'].iloc[-1] + newgreeks['TradeVega'].iloc[-1]
        tempTradeCol['VegaP&L'] = tradegreeks['TradeVegaPL'].iloc[-1]
        tempTradeCol['GreeksPL'] = tradegreeks['TradeGreeksPL'].iloc[-1]
        tempTradeCol['GreeksError'] = tradegreeks['TradePL'].iloc[-1] - tradegreeks['TradeGreeksPL'].iloc[-1]

        tempNewRisk = pd.Series(index=TempInfoSeri)
        tempNewRisk[self.TDate_yesterday] = newgreeks['OptionFee'].iloc[-1]
        tempNewRisk[self.TDate_today] = newgreeks['RiskPrice'].iloc[-1]
        tempNewRisk['Profit&Loss'] = newgreeks['OptionFee'].iloc[-1] + newgreeks['RiskPrice'].iloc[-1]
        tempNewRisk['BiasP&L'] = 0
        tempNewRisk['PositionP&L'] = 0
        tempNewRisk['DeltaP&L'] = 0
        tempNewRisk['GammaP&L'] = 0
        tempNewRisk['Theta'] = newgreeks['RiskTheta'].iloc[-1]
        tempNewRisk['ThetaP&L'] = 0
        tempNewRisk['Gamma+ThetaP&L'] = 0
        tempNewRisk['Vega'] = newgreeks['RiskVega'].iloc[-1]
        tempNewRisk['VegaP&L'] = 0
        tempNewRisk['GreeksPL'] = newgreeks['OptionFee'].iloc[-1] + newgreeks['RiskPrice'].iloc[-1]
        tempNewRisk['GreeksError'] = 0

        for i in range(self.UnderNum):
            deltastr = 'Delta' + str(i + 1)
            tempRiskCol[deltastr] = riskgreeks[deltastr].iloc[-1] + newgreeks['Risk' + deltastr].iloc[-1]
            tempTradeCol[deltastr] = tradegreeks[deltastr].iloc[-1] + newgreeks['Trade' + deltastr].iloc[-1]
            tempNewRisk[deltastr] = newgreeks['Risk' + deltastr].iloc[-1]
            gammastr = 'Gamma' + str(i + 1)
            tempRiskCol[gammastr] = riskgreeks[gammastr].iloc[-1] + newgreeks['Risk' + gammastr].iloc[-1]
            tempTradeCol[gammastr] = tradegreeks[gammastr].iloc[-1] + newgreeks['Trade' + gammastr].iloc[-1]
            tempNewRisk[gammastr] = newgreeks['Risk' + gammastr].iloc[-1]

        self.PoolInfo['SSigB'] = tempRiskCol
        self.PoolInfo['SSigT'] = tempTradeCol

        tempRiskTrade = tempRiskCol - tempTradeCol
        self.PoolInfo['Risk-Trade'] = tempRiskTrade

        tempRiskTotalPL = tempfuturepl + tempRiskCol
        tempRiskTotalPL['PositionP&L'] = tempRiskTotalPL['DeltaP&L'] - tempRiskTotalPL['BiasP&L']
        self.PoolInfo['TotalP&L'] = tempRiskTotalPL

        self.PoolInfo['NewP&L'] = tempNewRisk
        self.PoolInfo['TotalExistP&L'] = tempRiskTotalPL - tempNewRisk
        self.PoolInfo = np.round(self.PoolInfo)
        self.PoolInfo = self.PoolInfo.reindex(columns=PoolTitle)
        self.PoolInfo.to_excel(writer, sheet_name=u'交易盈亏分解', index=True)
        worksheet5 = writer.sheets[u'交易盈亏分解']
        worksheet5.set_column('A:S', 14, format2)
        worksheet5.set_column('F:I', 14, format3)
        worksheet5.conditional_format('H2:H' + str(1 + PoolInfoLen),
                                      {'type': 'cell', 'criteria': '==', 'value': 0, 'format': format4})
        writer.save()
        self.GatherExoticReport()

    def calgreeks(self, ParaType='Risk'):
        TableTitle = self.RiskTableTitle if ParaType == 'Risk' else self.TradeTableTitle
        TempPara1 = self.RiskPara1 if ParaType == 'Risk' else self.TradePara1
        TempPara2 = self.RiskPara2 if ParaType == 'Risk' else self.TradePara2
        exogreeks = pd.DataFrame()
        S_yesterday = self.PriceDict.loc[self.TDate_yesterday]
        S_today = self.PriceDict.loc[self.TDate_today]
        for i in range(np.size(self.opinfo, axis=0)):
            singleinfo = self.opinfo.iloc[i, :].copy()
            actualnotamt = (singleinfo.Notamt * singleinfo.Maturity /
                            singleinfo.YearDays * singleinfo.Pratio *
                            singleinfo.Pos if singleinfo.IfActual == 0 else
                            singleinfo.Notamt * singleinfo.Pratio *
                            singleinfo.Pos)
            multicontract = actualnotamt / singleinfo.InitialS
            t1 = of.mytdate(self.TDate_yesterday, singleinfo.EndDate)
            t2 = (of.mytdate(self.TDate_today, singleinfo.ActualEndDate)
                  if singleinfo.OptionType in AmericanType else
                  of.mytdate(self.TDate_today, singleinfo.EndDate))
            TempS0 = self.CorrS0[singleinfo.Code]
            tempMonte_yesterday = CorrOptionDict[singleinfo.OptionType](S_yesterday, t1, TempPara1, TempS0,
                                                                        'Price', singleinfo) * multicontract
            tempMonte_today = CorrOptionDict[singleinfo.OptionType](S_today, t2, TempPara2, TempS0, 'Price',
                                                                    singleinfo) * multicontract
            tempMontePL = np.array([tempMonte_yesterday, tempMonte_today, tempMonte_today - tempMonte_yesterday])
            tempPrice_yesterday = CorrOptionDict[singleinfo.OptionType](S_yesterday, t1, TempPara1, TempS0, 'Price',
                                                                        singleinfo) * multicontract
            tempPrice_today = CorrOptionDict[singleinfo.OptionType](S_today, t2, TempPara2, TempS0, 'Price',
                                                                    singleinfo) * multicontract
            tempPricePL = np.array([tempPrice_yesterday, tempPrice_today, tempPrice_today - tempPrice_yesterday])
            tempDelta_yesterday = CorrOptionDict[singleinfo.OptionType](S_yesterday, t1, TempPara1, TempS0, 'Delta',
                                                                        singleinfo) * multicontract
            tempDelta_today = CorrOptionDict[singleinfo.OptionType](S_today, t2, TempPara2, TempS0, 'Delta',
                                                                    singleinfo) * multicontract
            tempDeltaPL = np.append(tempDelta_today, tempDelta_yesterday.dot(S_today - S_yesterday))
            tempGamma_yesterday = CorrOptionDict[singleinfo.OptionType](S_yesterday, t1, TempPara1, TempS0, 'Gamma',
                                                                        singleinfo) * multicontract
            tempGamma_today = CorrOptionDict[singleinfo.OptionType](S_today, t2, TempPara2, TempS0, 'Gamma',
                                                                    singleinfo) * multicontract
            GammaPL = 0.5 * np.array(S_today - S_yesterday) * np.matrix(tempGamma_yesterday) * np.matrix(
                S_today - S_yesterday).T
            tempGamma_today = np.diag(tempGamma_today)
            tempGammaPL = np.append(tempGamma_today, GammaPL[0, 0])
            tempTheta_yesterday = CorrOptionDict[singleinfo.OptionType](S_yesterday, t1, TempPara1, TempS0, 'Theta',
                                                                        singleinfo) * multicontract
            tempTheta_today = CorrOptionDict[singleinfo.OptionType](S_today, t2, TempPara2, TempS0, 'Theta',
                                                                    singleinfo) * multicontract
            tempThetaPL = np.array([tempTheta_today, tempTheta_yesterday / YearDays])
            tempVega_yesterday = CorrOptionDict[singleinfo.OptionType](S_yesterday, t1, TempPara1, TempS0, 'Vega',
                                                                       singleinfo) * multicontract
            tempVega_today = CorrOptionDict[singleinfo.OptionType](S_today, t2, TempPara2, TempS0, 'Vega',
                                                                   singleinfo) * multicontract
            if np.isnan(TempPara2.loc['Rho'][0]):
                tempVegaPL = np.array(
                    [tempVega_today, tempVega_yesterday * (TempPara2.loc['Vol'][0] - TempPara1.loc['Vol'][0])])
            else:
                tempVegaPL = np.array([tempVega_today, tempVega_yesterday * (
                            CorrOptionDict[singleinfo.OptionType](S_today, t2, TempPara2, TempS0, 'Vol', singleinfo) -
                            CorrOptionDict[singleinfo.OptionType](S_yesterday, t1, TempPara1, TempS0, 'Vol',
                                                                  singleinfo))])
            tempGreeksPL = tempDeltaPL[-1] + tempGammaPL[-1] + tempThetaPL[-1] + tempVegaPL[-1]

            try:
                tempbiasPL = (0 if self.NotBias else tempDelta_yesterday.dot(S_today.values - S_yesterday.values -
                                                                             self.FutureDict.loc[
                                                                                 self.TDate_today].values +
                                                                             self.FutureDict.loc[
                                                                                 self.TDate_yesterday].values))
            except:
                print(self.NotBias, S_today.values, S_yesterday.values, self.FutureDict.loc[self.TDate_today].values,
                      self.FutureDict.loc[self.TDate_yesterday].values)
                print(self.UNCode)
            tempnum = np.concatenate([tempMontePL, tempPricePL, tempDeltaPL,
                                      tempGammaPL, tempThetaPL, tempVegaPL,
                                      [tempGreeksPL, tempbiasPL]]) * (-1)  # 按照现金流来计算
            tempgreeks = pd.DataFrame(tempnum, index=TableTitle[2:]).T
            tempgreeks[TableTitle[0]] = singleinfo.ContractId
            tempgreeks[TableTitle[1]] = singleinfo.Code
            exogreeks = exogreeks.append(tempgreeks)
        totaltemp = exogreeks.iloc[:, :-2].sum()
        totaltemp[TableTitle[0]] = 'sum'
        totaltemp[TableTitle[1]] = ''
        exogreeks = exogreeks.append(totaltemp, ignore_index=True)
        exogreeks = exogreeks.reindex(columns=TableTitle)
        return exogreeks

    def newcalgreeks(self):
        newgreeks = pd.DataFrame()
        S_today = self.PriceDict.loc[self.TDate_today]
        for j in range(np.size(self.newinfo, axis=0)):
            newsingleinfo = self.newinfo.iloc[j, :].copy()
            # 这里Basket产品的价格肯定有
            if newsingleinfo.ContractId == 'SWHY-GTJA-18001':
                """
                因這比期权合约是用結算价，与常规的收盘价不同
                """
                S_today_settle = 539.5
                TotalGreeks = np.concatenate([self.GetGreeks(S_today_settle, newsingleinfo, self.RiskPara2),
                                              self.GetGreeks(S_today_settle, newsingleinfo, self.TradePara2)]) * (-1)
                tempnewgreeks = pd.DataFrame(TotalGreeks, index=self.NewTableTitle[3:]).T

                tempfee = (newsingleinfo.Notamt * newsingleinfo.Maturity /
                           newsingleinfo.YearDays * newsingleinfo.Premium
                           if newsingleinfo.IfActual == 0 else
                           newsingleinfo.Notamt * newsingleinfo.Premium)
                tempnewgreeks[self.NewTableTitle[2]] = tempfee
                tempnewgreeks[self.NewTableTitle[0]] = newsingleinfo.ContractId
                tempnewgreeks[self.NewTableTitle[1]] = newsingleinfo.Code
                newgreeks = newgreeks.append(tempnewgreeks)
            else:
                TotalGreeks = np.concatenate([self.GetGreeks(S_today, newsingleinfo, self.RiskPara2),
                                              self.GetGreeks(S_today, newsingleinfo, self.TradePara2)]) * (-1)

                tempnewgreeks = pd.DataFrame(TotalGreeks, index=self.NewTableTitle[3:]).T

                tempfee = (newsingleinfo.Notamt * newsingleinfo.Maturity /
                           newsingleinfo.YearDays * newsingleinfo.Premium
                           if newsingleinfo.IfActual == 0 else
                           newsingleinfo.Notamt * newsingleinfo.Premium)
                tempnewgreeks[self.NewTableTitle[2]] = tempfee
                tempnewgreeks[self.NewTableTitle[0]] = newsingleinfo.ContractId
                tempnewgreeks[self.NewTableTitle[1]] = newsingleinfo.Code
                newgreeks = newgreeks.append(tempnewgreeks)
        totaltemp = newgreeks.iloc[:, :-2].sum()
        totaltemp[self.NewTableTitle[0]] = 'sum'
        totaltemp[self.NewTableTitle[1]] = ''

        newgreeks = newgreeks.append(totaltemp, ignore_index=True)
        newgreeks = newgreeks.reindex(columns=self.NewTableTitle)
        return newgreeks

    def GetGreeks(self, S, singleinfo, Para):
        multicontract = (singleinfo.Notamt * singleinfo.Maturity /
                         singleinfo.YearDays * singleinfo.Pratio * singleinfo.Pos
                         if singleinfo.IfActual == 0 else
                         singleinfo.Notamt * singleinfo.Pratio * singleinfo.Pos) / singleinfo.InitialS
        TempS0 = self.CorrS0[singleinfo.Code]
        t2 = of.mytdate(self.TDate_today, singleinfo.EndDate)

        tempMonte = CorrOptionDict[singleinfo.OptionType](S, t2, Para, TempS0, 'MontePrice', singleinfo) * multicontract
        tempPrice = CorrOptionDict[singleinfo.OptionType](S, t2, Para, TempS0, 'Price', singleinfo) * multicontract
        tempDelta = CorrOptionDict[singleinfo.OptionType](S, t2, Para, TempS0, 'Delta', singleinfo) * multicontract
        tempGamma = np.diag(
            CorrOptionDict[singleinfo.OptionType](S, t2, Para, TempS0, 'Gamma', singleinfo) * multicontract)
        tempTheta = CorrOptionDict[singleinfo.OptionType](S, t2, Para, TempS0, 'Theta', singleinfo) * multicontract
        tempVega = CorrOptionDict[singleinfo.OptionType](S, t2, Para, TempS0, 'Vega', singleinfo) * multicontract
        return np.concatenate([[tempMonte, tempPrice], tempDelta, tempGamma, [tempTheta, tempVega]])

    def GetTouchExpire(self):
        # touchcol = np.nonzero(self.opinfo.columns=='ActualEndDate')[0][0]+1
        exercisecol = np.nonzero(self.opinfo.columns == 'ExerciseFee')[0][0] + 1
        wb = openpyxl.load_workbook(self.infoname)
        # 因为版本的问题，这里的openpyxl必须是2.3.2版本，如果更高版本可能会出错，望留意
        ws = wb['Sheet1']
        for i in range(np.size(self.opinfo, axis=0)):
            singleinfo = self.opinfo.iloc[i, :].copy()
            tempEndDate = singleinfo.ActualEndDate
            if tempEndDate <= self.TDate_today:
                S = self.PriceDict.loc[self.TDate_today]
                multicontract = (singleinfo.Notamt * singleinfo.Maturity /
                                 singleinfo.YearDays * singleinfo.Pratio * singleinfo.Pos
                                 if singleinfo.IfActual == 0 else
                                 singleinfo.Notamt * singleinfo.Pratio * singleinfo.Pos) / singleinfo.InitialS
                exercisefee = CorrOptionDict[singleinfo.OptionType](S, 0, self.RiskPara2, self.CorrS0[singleinfo.Code],
                                                                    'Price', singleinfo) * multicontract
                ws.cell(row=singleinfo['ID'] + 1, column=exercisecol, value=exercisefee)
                self.allopinfo['ExerciseFee'][self.allopinfo['ID'] == singleinfo['ID']] = exercisefee
        wb.save(self.infoname)

    def GatherExoticReport(self):
        gatherfile = '.\ProfitLossAnalyst\OTC_PL_Decompose_Summary.xlsx'
        temppl = [self.TDate_today,
                  self.PoolInfo['TotalP&L']['BiasP&L'],
                  self.PoolInfo['TotalP&L']['PositionP&L'],
                  self.PoolInfo['TotalP&L']['DeltaP&L'],
                  self.PoolInfo['TotalP&L']['GammaP&L'],
                  self.PoolInfo['TotalP&L']['ThetaP&L'],
                  self.PoolInfo['TotalP&L']['Gamma+ThetaP&L'],
                  self.PoolInfo['TotalP&L']['VegaP&L'],
                  self.PoolInfo['TotalP&L']['GreeksPL'],
                  self.PoolInfo['SSigB']['GreeksError'],
                  self.PoolInfo['TotalExistP&L']['Profit&Loss'],
                  self.PoolInfo['NewP&L'][self.TDate_yesterday],
                  self.PoolInfo['NewP&L'][self.TDate_today],
                  self.PoolInfo['NewP&L']['Profit&Loss'],
                  self.PoolInfo['TotalP&L']['Profit&Loss']]
        wb = openpyxl.load_workbook(gatherfile)
        num_form = '_ * #,##0_ ;_ * \\-#,##0_ ;_ * "-"??_ ;_ @_ '
        try:
            ws = wb[self.UNCode]
            ifend = 1
            for i in range(1, len(list(ws.rows))):
                if ws.cell(row=i + 1, column=1).value >= self.TDate_today:
                    ifend = 0
                    break
            sindex = i + ifend + 1
            for tindex, tcell in enumerate(temppl):
                ws.cell(row=sindex, column=tindex + 1, value=tcell)
                ws.cell(row=sindex, column=tindex + 1).number_format = num_form
                if isinstance(tcell, datetime):
                    ws.cell(row=sindex, column=tindex + 1).number_format = 'mm-dd-yy'
                sumvalue = temppl[-1] + ws.cell(row=sindex - 1, column=len(temppl) + 1).value
                ws.cell(row=sindex, column=len(temppl) + 1, value=sumvalue)
                ws.cell(row=sindex, column=len(temppl) + 1).number_format = num_form
                wb.save(gatherfile)
        except:
            ws = wb.create_sheet(self.UNCode)
            ws.cell(row=1, column=1).value = 'TDate'
            ws.cell(row=1, column=2).value = 'BiasPL'
            ws.cell(row=1, column=3).value = 'PositionPL'
            ws.cell(row=1, column=4).value = 'DeltaPL'
            ws.cell(row=1, column=5).value = 'GammaPL'
            ws.cell(row=1, column=6).value = 'ThetaPL'
            ws.cell(row=1, column=7).value = 'GammaThetaPL'
            ws.cell(row=1, column=8).value = 'VegaPL'
            ws.cell(row=1, column=9).value = 'GreeksPL'
            ws.cell(row=1, column=10).value = 'GreeksError'
            ws.cell(row=1, column=11).value = 'TotalExistPL'
            ws.cell(row=1, column=12).value = 'NewFee'
            ws.cell(row=1, column=13).value = 'NewValue'
            ws.cell(row=1, column=14).value = 'NewPL'
            ws.cell(row=1, column=15).value = 'TotalPL'
            ws.cell(row=1, column=16).value = 'SumPL'
            sindex = 2
            for tindex, tcell in enumerate(temppl):
                ws.cell(row=sindex, column=tindex + 1, value=tcell)
                ws.cell(row=sindex, column=tindex + 1).number_format = num_form
                if isinstance(tcell, datetime):
                    ws.cell(row=sindex, column=tindex + 1).number_format = 'mm-dd-yy'
                sumvalue = temppl[-1]
                ws.cell(row=sindex, column=len(temppl) + 1, value=sumvalue)
                ws.cell(row=sindex, column=len(temppl) + 1).number_format = num_form
                wb.save(gatherfile)


if __name__ == '__main__':
    Report = ExoticReport('Basket0', datetime(2016,12,16))

